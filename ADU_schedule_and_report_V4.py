import os
import zipfile
import csv
import shutil

import pandas as pd
import numpy as np
import math
from collections import defaultdict

import time
from dateutil.parser import parse
import datetime as dt
from datetime import datetime

import xlsxwriter
from pyxlsb import open_workbook as open_xlsb
import win32com.client
import xlwings as xw

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.utils import get_column_letter




# Global parameter
DIR_INPUT  ='//ion.media/files/APPS/Analytics/_Data_/Misc/ADU Trust 3.0/adu_raw_data/'
#DIR_INPUT ='C:/ION/Commercial/ADU_Report/V2/'
DIR_OUTPUT='//ion.media/files/APPS/Analytics/_Data_/Misc/ADU Trust 3.0/adu_reports/'
#DIR_OUTPUT ='//ion.media/files/APPS/Analytics/_Data_/Misc/ADU Trust 3.0/adu_test/'
DIR_ARCHIVE='//ion.media/files/APPS/Analytics/_Data_/Misc/ADU Trust 3.0/adu_raw_data/history_raw/'
DIR_REPORT ='//ion.media/shared/1 Commercial/! IM 3.0/2 Control 2/'
#DIR_REPORT='C:/ION/Commercial/ADU_Report/V2/'
P = set(['Holiday Movies (Prime)', 'ION Originals (Prime)', 'Prime', 'Prime no CM',"Valentine's Day Movie (Prime)"])
NP = set(['Daytime (M-F)', 'Early Morning (M-S)', 'Fringe (M-S)', 'Holiday Movies (Non Prime)', \
          'Late Night (M-S)', 'Morning (M-S)', 'Non-Prime ROS**', 'Non-Prime ROS*', 'Weekend Day (S-Sun)',\
          'Non-Prime (M-Su 11A-7P)','Non-Prime (M-Su 1A-3A)','Non-Prime (M-Su 7A-11A)','Non-Prime'])

ADU_Pmix_low = 0.25
ADU_Pmix_high = 0.95
ADU_Pmix_shift = 0.0

delv_perc_bar = 0.99
more_units_perc = 0.0

# Compare dates
def date_comparison(date1, date2):
    date1 = parse(str(date1))
    date2 = parse(str(date2))
    return date1 < date2

# Number of weeks between two dates. 0619-0625 has 0 week in between, 0619-0626 has 1 week in between
def weeks_between(d1, d2):
    d1 = parse(str(d1))
    d2 = parse(str(d2))
    return math.trunc((d2 - d1).days / 7)


# transform different selling titles to P and NP
def dayparts(r):
    global P
    global NP
    if r['Selling Title'] in P:
        return 'P'
    elif r['Selling Title'] in NP:
        return 'NP'
    else: 
        print('Missing Daypart, Please check data: ',r['Selling Title'])
    return None


# structured class to store each guarantee ID group
class GID:
    def __init__(self, r):
        self.row = r
        self.DealName = set()
        self.DealNum = set()
        self.AEName = set()
        self.Agency = set()
        self.Marketplace = set()
        dictionary = {'Booked $': 0, 'Deal Imp': 0, 'Delv Imp': 0, 'Imps Owed': 0, 'Units': 0, 'CPM': 0}
        self.Sold_P = dictionary.copy()
        self.Sold_NP = dictionary.copy()
        self.ADU_P = dictionary.copy()
        self.ADU_NP = dictionary.copy()
        self.Total = dictionary.copy()
        self.P = {'Guar': 0, 'Est': 0, 'Delv': 0, 'Forecast Imp': 0, 'ADUs': 0}
        self.NP = self.P.copy()

    def new_info(self, ratings):
        if len(str(self.row['Guarantee Name']))<4:
            self.GName = self.row['Deal Name']
        else:
            self.GName = self.row['Guarantee Name']
        self.DealNum.add(self.row['Deal Numbers in Guarantee'])
        self.Marketplace.add(self.row['Marketplace'])
        self.Advertiser = self.row['Advertiser']
        self.AEName.add(self.row['AE Name'])
        self.Agency.add(self.row['Agency Name (Billing)'])
        self.DealName.add(self.row['Deal Name'])
        self.SoldDemo = self.row['Primary Demo']
        self.StartDate = self.row['Week Start Date']
        self.EndDate = self.row['Week End Date']
        self.DealYear = self.row['Deal Year']

        # forecast ratings
        try:
            self.P['Forecast Imp'] = ratings.loc[ratings['Demo'] == self.SoldDemo, 'Prime Imp'].iloc[0]
            self.NP['Forecast Imp'] = ratings.loc[ratings['Demo'] == self.SoldDemo, 'Non Prime Imp'].iloc[0]
        except:
            print(self.SoldDemo)
        self.update_by_daypart(self.row)
        return


    def update_info(self, r):
        self.DealName.add(r['Deal Name'])
        self.Marketplace.add(r['Marketplace'])
        self.AEName.add(r['AE Name'])
        self.Agency.add(r['Agency Name (Billing)'])

        if not date_comparison(self.StartDate, r['Week Start Date']):
            self.StartDate = r['Week Start Date']
        if date_comparison(self.EndDate, r['Week End Date']):
            self.EndDate = r['Week End Date']
        self.update_by_daypart(r)



    def update_by_daypart(self, r):
        if r['ADU Ind'] == 'N':
            if dayparts(r) == 'P':
                self.Sold_P['Booked $'] += r['Booked Dollars']
                self.Sold_P['Deal Imp'] += r['Primary Demo Non-ADU Equiv Deal Imp'] / 1000
                self.Sold_P['Delv Imp'] += r['Primary Demo Equiv Post Imp'] / 1000
                self.Sold_P['Imps Owed'] = self.Sold_P['Deal Imp'] - self.Sold_P['Delv Imp']
                self.Sold_P['Units'] += r['Equiv Units']
                self.Sold_P['CPM'] = self.Sold_P['Booked $'] / self.Sold_P['Deal Imp'] if self.Sold_P['Deal Imp'] else 0
            elif dayparts(r) == 'NP':
                self.Sold_NP['Booked $'] += r['Booked Dollars']
                self.Sold_NP['Deal Imp'] += r['Primary Demo Non-ADU Equiv Deal Imp'] / 1000
                self.Sold_NP['Delv Imp'] += r['Primary Demo Equiv Post Imp'] / 1000
                self.Sold_NP['Imps Owed'] = self.Sold_NP['Deal Imp'] - self.Sold_NP['Delv Imp']
                self.Sold_NP['Units'] += r['Equiv Units']
                self.Sold_NP['CPM'] = self.Sold_NP['Booked $'] / self.Sold_NP['Deal Imp'] if self.Sold_NP['Deal Imp'] else 0

        else:
            if dayparts(r) == 'P':
                self.ADU_P['Delv Imp'] += r['Primary Demo Equiv Post Imp'] / 1000
                self.ADU_P['Imps Owed'] = self.ADU_P['Deal Imp'] - self.ADU_P['Delv Imp']
                self.ADU_P['Units'] += r['Equiv Units']

            elif dayparts(r) == 'NP':
                self.ADU_NP['Delv Imp'] += r['Primary Demo Equiv Post Imp'] / 1000
                self.ADU_NP['Imps Owed'] = self.ADU_NP['Deal Imp'] - self.ADU_NP['Delv Imp']
                self.ADU_NP['Units'] += r['Equiv Units']

        self.Total['Booked $'] = self.Sold_P['Booked $'] + self.Sold_NP['Booked $']
        self.Total['Deal Imp'] = self.Sold_P['Deal Imp'] + self.Sold_NP['Deal Imp']
        self.Total['Delv Imp'] = self.Sold_P['Delv Imp'] + self.Sold_NP['Delv Imp'] \
                                 + self.ADU_P['Delv Imp'] + self.ADU_NP['Delv Imp']
        self.Total['% Delv'] = self.Total['Delv Imp'] / self.Total['Deal Imp'] if self.Total['Deal Imp'] else 0
        self.Total['Imps Owed'] = self.Total['Deal Imp'] - self.Total['Delv Imp']
        self.Total['Units'] = self.Sold_P['Units'] + self.Sold_NP['Units'] + self.ADU_P['Units'] + self.ADU_NP['Units']
        self.Total['CPM'] = self.Total['Booked $'] / self.Total['Deal Imp'] if self.Total['Deal Imp'] else 0
        self.Total['Liability $'] = max(0, self.Total['Imps Owed'] * self.Total['CPM'])
        self.Total['P Mix %'] = self.Sold_P['Deal Imp'] / self.Total['Deal Imp'] if self.Total['Deal Imp'] else 0
        self.Total['NP Mix %'] = 1 - self.Total['P Mix %']

        self.Adj_P_ADU, self.Adj_NP_ADU = Adjust_ADU_P_Mixture(self.Total['P Mix %'])
        NP_more_units_perc =1.0
        if self.Total['% Delv'] >= delv_perc_bar:
            NP_more_units_perc = 1.0+more_units_perc
          
        
        self.P['Guar'] = self.Sold_P['Deal Imp'] / self.Sold_P['Units'] if self.Sold_P['Units'] else 0
        self.P['ADUs'] = round(self.Adj_P_ADU * self.Total['Imps Owed'] / self.P['Forecast Imp'])
        self.P['Est'] = self.Sold_P['Delv Imp'] / self.Sold_P['Units'] if self.Sold_P['Units'] else 0
        self.P['Delv'] = self.P['Est'] / self.P['Guar'] if self.P['Guar'] else 0

        self.NP['Guar'] = self.Sold_NP['Deal Imp'] / self.Sold_NP['Units'] if self.Sold_NP['Units'] else 0
        self.NP['ADUs'] =max(0,round((self.Total['Imps Owed'] - self.P['ADUs']*self.P['Forecast Imp']) / self.NP['Forecast Imp']*NP_more_units_perc))
        self.NP['Est'] = self.Sold_NP['Delv Imp'] / self.Sold_NP['Units'] if self.Sold_NP['Units'] else 0
        self.NP['Delv'] = self.NP['Est'] / self.NP['Guar'] if self.NP['Guar'] else 0
        self.Total['ADUs'] = self.P['ADUs'] + self.NP['ADUs']

        return 

    def GID_to_List(self):
        return [self.GName, self.DealNum, self.Marketplace, self.Advertiser, \
             self.AEName, self.Agency, self.DealName, self.SoldDemo, self.StartDate, self.EndDate, self.DealYear] \
             + [self.Sold_P, self.Sold_NP, self.ADU_P, self.ADU_NP, self.Total, self.P, self.NP]

    
def Adjust_ADU_P_Mixture(P_mix):
    if P_mix >= ADU_Pmix_high or P_mix <= ADU_Pmix_low:
        adj_P_mix = P_mix
    else:
        adj_P_mix = max(ADU_Pmix_low, P_mix - ADU_Pmix_shift)
    adj_NP_mix = 1-adj_P_mix
           
    return adj_P_mix, adj_NP_mix


def get_dict(df, ratings, endq):
    dic = dict()
    for col, row in df.iterrows():
        if date_comparison(row['Week Start Date'], endq):
            # if the Guarantee ID has not shown before, get new info
            if row['Guarantee ID'] not in dic: 
                c = GID(row)
                c.new_info(ratings)
                dic[row['Guarantee ID']] = c
            # else update the information
            else: 
                dic[row['Guarantee ID']].update_info(row)
        else:
            continue
    return dic


# read in the result from get_dict, turn dictionary to dataframe of all guarantee id deal
def form_df(result):
    column_names = ['Guarantee ID', 'Guarantee Name', 'Deal ID', 'Marketplace',\
                    'Advertiser', \
                    'AE Name', 'Agency', 'Deal Name', 'Primary Demo', 'Sold Start Date', 'Sold End Date', 'Deal Year',\
                    'Sold Prime Booked $', 'Sold Prime Deal Imp', 'Sold Prime Delv Imp', 'Sold Prime Imps Owed',
                    'Sold Prime Units', \
                    'Sold Prime CPM', 'Sold NP Booked $', 'Sold NP Deal Imp', 'Sold NP Delv Imp', 'Sold NP Imps Owed', \
                    'Sold NP Units', 'Sold NP CPM', 'ADU Prime Booked $', 'ADU Prime Deal Imp', 'ADU Prime Delv Imp', \
                    'ADU Prime Imps Owed', 'ADU Prime Units', 'ADU Prime CPM', 'ADU NP Booked $', 'ADU NP Deal Imp', \
                    'ADU NP Delv Imp', 'ADU NP Imps Owed', 'ADU NP Units', 'ADU NP CPM', 'Total Booked $', \
                    'Total Deal Imp', 'Total Delv Imp', 'Total Imps Owed', 'Total Units', 'Total CPM', \
                    'Total % Delv', 'Total Liability $', 'Total P Mix %', 'Total NP Mix %', 'Total ADUs', \
                    'P Guar', 'P Est', 'P Delv', 'P Forecast Imp', 'P ADUs', 'NP Guar', 'NP Est', 'NP Delv', 'NP Forecast Imp', \
                    'NP ADUs']
    rows = []
    for k, v in result.items():
        row = []
        row.append(k)  # G_ID

        for element in v.GID_to_List():
            if type(element) is set:
                element = list(element)
                s = ','.join(str(e) for e in element)
                if len(s)>200:
                    s = s[:200] + '...'
                row.append(s)
            elif type(element) is dict:
                if len(element) == 6:
                    for key in ['Booked $', 'Deal Imp', 'Delv Imp', 'Imps Owed', 'Units', 'CPM']:
                        row.append(element[key])
                elif len(element) ==5:
                    for key in ['Guar', 'Est', 'Delv', 'Forecast Imp', 'ADUs']:
                        row.append(element[key])
                else:
                    for key in ['Booked $', 'Deal Imp', 'Delv Imp', 'Imps Owed', 'Units', 'CPM', \
                        '% Delv', 'Liability $', 'P Mix %', 'NP Mix %', 'ADUs']:
                        row.append(element[key])
            else:
                row.append(element)
        
        rows.append(row)
    output = pd.DataFrame(rows)
    output.columns = column_names
    output = output[['Deal ID', 'Guarantee ID', 'Guarantee Name', 'Marketplace', \
                     'Advertiser', \
                     'AE Name', 'Agency', 'Deal Name','Deal Year', 'Primary Demo', 'Sold Start Date', 'Sold End Date',  \
                     'Sold Prime Booked $', 'Sold Prime Deal Imp', 'Sold Prime Delv Imp', 'Sold Prime Imps Owed',
                     'Sold Prime Units', \
                     'Sold Prime CPM', 'Sold NP Booked $', 'Sold NP Deal Imp', 'Sold NP Delv Imp', 'Sold NP Imps Owed', \
                     'Sold NP Units', 'Sold NP CPM', 'ADU Prime Booked $', 'ADU Prime Deal Imp', 'ADU Prime Delv Imp', \
                     'ADU Prime Imps Owed', 'ADU Prime Units', 'ADU Prime CPM', 'ADU NP Booked $', 'ADU NP Deal Imp', \
                     'ADU NP Delv Imp', 'ADU NP Imps Owed', 'ADU NP Units', 'ADU NP CPM', 'Total Booked $', \
                     'Total Deal Imp', 'Total Delv Imp', 'Total % Delv', 'Total Imps Owed', 'Total Units', 'Total CPM', \
                     'Total Liability $', 'Total P Mix %', 'Total NP Mix %', 'P Guar', \
                     'NP Guar', 'P Est', 'NP Est', 'P Delv', 'NP Delv', 'P Forecast Imp', 'NP Forecast Imp', 'P ADUs', 'NP ADUs', \
                     'Total ADUs']]
    return output


# Mondays between two dates, excluding the last Monday. Used for scheduling ADU
def week_range(startdate, enddate):
    weeks = pd.date_range(startdate, enddate, freq='W-MON').strftime('%m/%d/%Y').tolist()
    return weeks[:-1]


# Find the previous, current, next and the next after next quarter giving the input date
def find_quarters(quarters, startdate):
    mon = pd.date_range(startdate, periods=1, freq='W-MON').strftime('%m/%d/%Y').tolist()[0]
    current_q = quarters.loc[quarters['start_date'].astype(str) == mon, 'quarter'].iloc[0][1]
    current_year = startdate.year
    current = (current_q, current_year)

    prev = (str(int(current_q) - 1), current_year)
    one_after = (str(int(current_q) + 1), current_year)
    two_after = (str(int(current_q) + 2), current_year)

    if current_q == '1':
        prev = ('4', current_year - 1)
    elif current_q == '4':
        one_after = ('1', current_year + 1)
        two_after = ('2', current_year + 1)
    elif current_q == '3':
        two_after = ('1', current_year + 1)

    return prev, current, one_after, two_after


# Find the start date of each quarters
def quarter_startdate(quarters, four_q):
    l = []
    for q in four_q:
        a = quarters[quarters['end_date'].str.contains(str(q[1]))]
        a = a[a['quarter'].astype(str) == 'Q' + q[0]]
        l.append(a['start_date'].iloc[0])
    return l


# Get the prime, nonprime and total baselayers
def past(df, startdate, enddate):
    parts = df[['Guarantee ID', 'Week Start Date', 'ADU Ind', 'Equiv Units', 'Selling Title']]
    weeks = week_range(startdate, enddate)

    dic_s_p = dict()
    dic_s_np = dict()
    dic_adu_p = dict()
    dic_adu_np = dict()
    dic_total_p = dict()
    dic_total_np = dict()

    for ind, row in parts.iterrows():
        if row['Guarantee ID'] not in dic_s_p:
            dic_s_p[row['Guarantee ID']] = [0] * len(weeks)
            dic_s_np[row['Guarantee ID']] = [0] * len(weeks)
            dic_adu_p[row['Guarantee ID']] = [0] * len(weeks)
            dic_adu_np[row['Guarantee ID']] = [0] * len(weeks)
            dic_total_p[row['Guarantee ID']] = [0] * len(weeks)
            dic_total_np[row['Guarantee ID']] = [0] * len(weeks)      
            
        #reformat_w = str(datetime.strptime(str(row['Week Start Date']).split()[0], '%Y-%m-%d').strftime('%m/%d/%Y'))
        reformat_w = row['Week Start Date']
        if reformat_w in weeks:
            if dayparts(row) == 'P':
                if row['ADU Ind'] == 'N':  # Prime spots
                    dic_s_p[row['Guarantee ID']][weeks.index(reformat_w)] += row['Equiv Units']
                else: # Prime ADU
                    dic_adu_p[row['Guarantee ID']][weeks.index(reformat_w)] += row['Equiv Units']
                dic_total_p[row['Guarantee ID']][weeks.index(reformat_w)] += row['Equiv Units'] # Prime Total
            else:
                if row['ADU Ind'] == 'N': # Nonprime spots
                    dic_s_np[row['Guarantee ID']][weeks.index(reformat_w)] += row['Equiv Units']
                else: # Nonprime ADU
                    dic_adu_np[row['Guarantee ID']][weeks.index(reformat_w)] += row['Equiv Units']
                dic_total_np[row['Guarantee ID']][weeks.index(reformat_w)] += row['Equiv Units'] #Nonprime Total
    return dic_s_p, dic_s_np, dic_adu_p, dic_adu_np, dic_total_p, dic_total_np


# round units to a multiple of 1, and calculate the leftover value
def round_unit(num):
    result = round(num)
    left = num - result
    return (result, left)

#@past_..  the sold/adu units in the spots  , called base line
#@df1  get_dict->form_df->output->df1 contain guarantee deal info
#@stq,edq, start/end quater
def schedule_ADU(past_s_p, past_adu_p, past_s_np, past_adu_np, df1, startq, endq, startdate):
    weeks = week_range(startq, endq)

    dic_p = dict()
    dic_np = dict()
    total_weeks = week_range(startq, endq)

    l = []
    for ind, row in df1.iterrows():
        # Schedule start date
        if date_comparison(row['Sold Start Date'], startdate):
            s = pd.date_range(startdate, periods=1, freq='W-MON').strftime('%m/%d/%Y').tolist()[0]
        else:
            s = row['Sold Start Date']
            s = pd.date_range(s, periods=1, freq='W-MON').strftime('%m/%d/%Y').tolist()[0]
        
        # Schedule end date
        if date_comparison(row['Sold End Date'], weeks[-1]):
            e = row['Sold End Date']
            e = pd.date_range(e, periods=1, freq='W-MON').strftime('%m/%d/%Y').tolist()[0]
        else:
            e = weeks[-1]
        weeks_left = weeks_between(s, e) + 1

        if weeks_left <= 0: # If no available week, do not schedule anything
            dic_p[row['Guarantee ID']] = [0] * len(weeks)
            dic_np[row['Guarantee ID']] = [0] * len(weeks)

        else:
            if row['Guarantee ID'] not in dic_p:
                dic_p[row['Guarantee ID']] = [0] * len(weeks)
                dic_np[row['Guarantee ID']] = [0] * len(weeks)
            
            if row['P ADUs'] > 0: #If there is Prime ADU
                scheduled_spots = past_s_p[row['Guarantee ID']][total_weeks.index(s):] # scheduled spots = prime spots baselayer
                total = sum(scheduled_spots[:total_weeks.index(weeks[-1]) + 1]) # total number of prime spots 
                new = dic_p[row['Guarantee ID']]
                if total == 0: # if no prime spots
                    try: # Check whether there is ADU scheduled
                        scheduled_spots = past_adu_p[row['Guarantee ID']][total_weeks.index(s):] # scheduled spots = prime ADU baselayer
                        total = sum(scheduled_spots[:total_weeks.index(weeks[-1]) + 1]) # total number of prime ADU
                    except: # if no spots and ADUs scheduled, do not schedule new ADU
                        dic_p[row['Guarantee ID']] = new
                        
                left = 0 
                for i in range(weeks.index(s), weeks.index(e) + 1): # schedule new ADU proportional to the scheduled_spots
                    if scheduled_spots[i - weeks.index(s)] != 0:
                        new[i] = round_unit(row['P ADUs'] / total * scheduled_spots[i - weeks.index(s)] + left)[0]
                        left = round_unit(row['P ADUs'] / total * scheduled_spots[i - weeks.index(s)] + left)[1]
                dic_p[row['Guarantee ID']] = new

            if row['NP ADUs'] > 0: #If there is Nonprime ADU

                scheduled_spots = past_s_np[row['Guarantee ID']][total_weeks.index(s):]
                total = sum(scheduled_spots[:total_weeks.index(weeks[-1]) + 1])
                new = dic_np[row['Guarantee ID']]
                if total == 0:
                    try:
                        scheduled_spots = past_adu_np[row['Guarantee ID']][total_weeks.index(s):]
                        total = sum(scheduled_spots[:total_weeks.index(weeks[-1]) + 1])
                    except:
                        dic_np[row['Guarantee ID']] = new
                        
                left = 0
                for i in range(weeks.index(s), weeks.index(e) + 1):
                    if scheduled_spots[i - weeks.index(s)] != 0:
                        new[i] = round_unit(row['NP ADUs'] / total * scheduled_spots[i - weeks.index(s)] + left)[0]
                        left = round_unit(row['NP ADUs'] / total * scheduled_spots[i - weeks.index(s)] + left)[1]
                dic_np[row['Guarantee ID']] = new        
                
                
                
            if row['P ADUs'] < 0: #If there is Prime ADU need to take back
                new = dic_p[row['Guarantee ID']]
                try: # Check whether there is ADU scheduled
                    scheduled_spots = past_adu_p[row['Guarantee ID']][total_weeks.index(s):] # scheduled spots = prime ADU baselayer
                    total = sum(scheduled_spots[:total_weeks.index(weeks[-1]) + 1]) # total number of prime ADU
                except: # if ADUs scheduled, do not take back ADU
                    dic_p[row['Guarantee ID']] = new
                        
                left = 0 
                for i in range(weeks.index(s), weeks.index(e) + 1): # take back ADU proportional to the scheduled_spots
                    u = scheduled_spots[i - weeks.index(s)]
                    if u != 0:
                        new[i] = -min(u, -round(row['P ADUs'] / total * u + left))
                        left = (row['P ADUs'] / total * u + left) - new[i]
                dic_p[row['Guarantee ID']] = new

            if row['NP ADUs'] < 0: #If there is Prime ADU need to take back
                new = dic_np[row['Guarantee ID']]
                try: # Check whether there is ADU scheduled
                    scheduled_spots = past_adu_np[row['Guarantee ID']][total_weeks.index(s):] # scheduled spots = prime ADU baselayer
                    total = sum(scheduled_spots[:total_weeks.index(weeks[-1]) + 1]) # total number of prime ADU
                except: # if ADUs scheduled, do not take back ADU
                    dic_np[row['Guarantee ID']] = new
                        
                left = 0 
                for i in range(weeks.index(s), weeks.index(e) + 1): # take back ADU proportional to the scheduled_spots
                    u = scheduled_spots[i - weeks.index(s)]
                    if u != 0:
                        new[i] = -min(u, -round(row['NP ADUs'] / total * u + left))
                        left = (row['NP ADUs'] / total * u + left) - new[i]
                dic_np[row['Guarantee ID']] = new
      
                
    schedule_p_df = pd.DataFrame.from_dict(dic_p, orient='index', columns=weeks).reset_index()
    schedule_p_df = schedule_p_df.rename(columns={'index': 'Guarantee ID'})
    schedule_np_df = pd.DataFrame.from_dict(dic_np, orient='index', columns=weeks).reset_index()
    schedule_np_df = schedule_np_df.rename(columns={'index': 'Guarantee ID'})
    return schedule_p_df, schedule_np_df


#@df is the raw dealmake data
#data_String: the data to start adu schdul
def raw_result(df, quarters, date_string, startdate, ratings, four_q, startq, endq):
   
    weeks = week_range(startq, endq)
    
    result = get_dict(df, ratings, endq)
    output = form_df(result)

    df1 = output[['Guarantee ID', 'Sold Start Date', 'Sold End Date', 'P ADUs', 'NP ADUs', 'Total ADUs']]

    past_s_p, past_s_np, past_adu_p, past_adu_np, bp, bnp = past(df, startq, endq)

    baselayer_p = pd.DataFrame.from_dict(bp, orient='index', columns=weeks).reset_index()
    baselayer_np = pd.DataFrame.from_dict(bnp, orient='index', columns=weeks).reset_index()
    
    baselayer_p = baselayer_p.rename(columns={'index': 'Guarantee ID'})
    baselayer_np = baselayer_np.rename(columns={'index': 'Guarantee ID'})
    baselayer_p.sort_values(by=['Guarantee ID'],inplace=True)
    baselayer_np.sort_values(by=['Guarantee ID'],inplace=True)

    P_ADU_schedule, NP_ADU_schedule = schedule_ADU(past_s_p, past_adu_p, past_s_np, past_adu_np, df1, startq, endq, startdate)
    P_ADU_schedule.sort_values(by=['Guarantee ID'],inplace=True)
    NP_ADU_schedule.sort_values(by=['Guarantee ID'],inplace=True)

    basic_info = output.copy()
    basic_info.sort_values(by=['Guarantee ID'],inplace=True)

    return date_string, basic_info, baselayer_p, baselayer_np, P_ADU_schedule, NP_ADU_schedule #, changeDF


def format_df(raw, new):
    writer = pd.ExcelWriter(DIR_OUTPUT+ 'YM -- 1 ION ADU 3.0 (Arjun) -- ' +str(datetime.now().strftime("%Y-%m-%d"))+'.xlsx', engine='xlsxwriter')
    workbook = writer.book
    
    # Set Font 
    workbook.formats[0].set_font_name('Arial')

    count_row = raw[1].shape[0] + 1  # gives number of row count
    count_col = raw[1].shape[1] + 3  # gives number of col count
    raw[1].to_excel(writer, sheet_name='ADUs to schedule', startrow=7, startcol=2, header=False, index = False)
    
    new.to_excel(DIR_OUTPUT+str(datetime.now().strftime('%Y-%m-%d'))+' ADU Data.xlsx',sheet_name='Data', index = False)

    worksheet = writer.sheets['ADUs to schedule']
    worksheet.set_zoom(85)

    # Clean the headers
    for col_num, value in enumerate(raw[1].columns.values):
        if col_num <= 9:
            worksheet.write(5, col_num + 2, value)
        elif col_num <= 11:
            worksheet.write(5, col_num + 2, ' '.join(value.split()[1:]))
        elif col_num <= 35:
            worksheet.write(5, col_num + 2, ' '.join(value.split()[2:]))
        else:
            worksheet.write(5, col_num + 2, ' '.join(value.split()[1:]))

    s = [] # stores the start column of each dataframe
    e = [] # stores the end column of each dataframe
    date_fmt = workbook.add_format({'num_format':'mm/dd/yyyy',
                                    'font_name': 'Arial'})     
    for i in range(2, len(raw)):
        raw[i].iloc[:, 1:].to_excel(writer, sheet_name='ADUs to schedule', startrow=7, startcol=count_col, index=False,
                                    header=False)
        if i < 4:
            for col_num, value in enumerate(raw[i].columns.values[1:]):
                worksheet.write(5, count_col + col_num, value)
        else:
            for col_num, value in enumerate(raw[i].columns.values[1:]):
                value = datetime.strptime(value, '%m/%d/%Y')
                worksheet.write_datetime(5, count_col + col_num, value, date_fmt)
        s.append(count_col)
        for r in range(count_row):
            for c in range(count_col, count_col):
                worksheet.write_blank(r, c, None)

        count_col += raw[i].shape[1]
        e.append(count_col - 2)
    s_letter = ['B'] #start column letter of each dataframe
    e_letter = ['L'] #end column letter of each dataframe
    for i in range(len(s)):
        s_letter.append(xlsxwriter.utility.xl_col_to_name(s[i]))
        e_letter.append(xlsxwriter.utility.xl_col_to_name(e[i]))

    # sum of scheduled spots and ADUs
    for i in range(len(e)):
        col = xlsxwriter.utility.xl_col_to_name(e[i] + 1)
        for r in range(8, count_row + 7):
            worksheet.write_formula(col + str(r),
                                    '{=SUM(' + s_letter[i + 1] + str(r) + ':' + e_letter[i + 1] + str(r) + ')}')
   
    # Deals not in flight
    col = xlsxwriter.utility.xl_col_to_name(e[-1] + 2)
    Total_P_ADU_col = xlsxwriter.utility.xl_col_to_name(e[2] + 1)
    Total_NP_ADU_col = xlsxwriter.utility.xl_col_to_name(e[3] + 1)
    Total_ADU_col = xlsxwriter.utility.xl_col_to_name(s[0] - 2)
    for r in range(8, count_row + 7):
        worksheet.write_formula(col + str(r), '{=' + Total_ADU_col + str(r) + '-' + Total_P_ADU_col + str(
            r) + '-' + Total_NP_ADU_col + str(r) + '}')

    # Header
    bold = workbook.add_format({'bold': True, 
                                'font_name': 'Arial'})
    worksheet.write(1, 2, 'ION Media', bold)
    worksheet.write(2, 2, 'ADU Trust 3.0', bold)
    bold_blue = workbook.add_format({'bold': True, 'font_color': 'blue', 
                                   'font_name': 'Arial'})
    white = workbook.add_format({'font_color': 'white', 
                                    'font_name': 'Arial'})
    worksheet.write(0, 2, str(datetime.now().strftime("%m/%d/%Y")), bold_blue)
    worksheet.write(0, 3, raw[0], white)

    
    # Add Title & Merge
    format_b = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#99CCFF', 
        'font_name': 'Arial'})
    format_o = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#FFCC99',
        'font_name': 'Arial'})
    format_y = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#FFFFCC',
        'font_name': 'Arial'})
    format_g = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#C0C0C0',
        'font_name': 'Arial'})  # grey

    try:
        #worksheet.merge_range('C4:I4', 'DEAL', format_g)
        #worksheet.merge_range('C5:I5', ' ', format_g)
        worksheet.merge_range(s_letter[3] + '4:' + e_letter[3] + '4', 'Prime - ADU Suggested Flighting', format_b)
        worksheet.merge_range(s_letter[4] + '4:' + e_letter[4] + '4', 'Non Prime - ADU Suggested Flighting', format_o)
        worksheet.merge_range(s_letter[1] + '4:' + e_letter[1] + '4', 'Prime Fligting - Sold Units', format_b)
        worksheet.merge_range(s_letter[2] + '4:' + e_letter[2] + '4', 'Non Prime Fligting - Sold Units', format_o)

    except:
        print('nope')


    # Headers for dataframes
    for i in range(2, 59):
        if i <= 13:
            worksheet.write(3, i, 'SOLD', format_g)
            worksheet.write(4, i, ' ', format_g)
        elif i <= 19:
            worksheet.write(3, i, 'SOLD', format_b)
            worksheet.write(4, i, 'Prime', format_b)
        elif i <= 25:
            worksheet.write(3, i, 'SOLD', format_o)
            worksheet.write(4, i, 'NP', format_o)
        elif i <= 31:
            worksheet.write(3, i, 'ADU', format_b)
            worksheet.write(4, i, 'Prime', format_b)
        elif i <= 37:
            worksheet.write(3, i, 'ADU', format_o)
            worksheet.write(4, i, 'NP', format_o)
        elif i <= 47:
            worksheet.write(3, i, 'Total', format_g)
            worksheet.write(4, i, ' ', format_g)
        else:
            worksheet.write(3, i, ' ', format_g)
            if i != 58:
                if i % 2 == 0:
                    worksheet.write(4, i, 'P', format_g)
                else:
                    worksheet.write(4, i, 'NP', format_g)
            else:
                worksheet.write(4, i, 'Total', format_g)

    for i in range(s[0], e[0] + 2):
        worksheet.write(4, i, 'P', format_b)
        if i == e[0] + 1:
            worksheet.write(3, i, 'Total', format_b)
    for i in range(s[1], e[1] + 2):
        worksheet.write(4, i, 'NP', format_o)
        if i == e[1] + 1:
            worksheet.write(3, i, 'Total', format_o)
    for i in range(s[2], e[2] + 2):
        worksheet.write(4, i, 'P', format_b)
        if i == e[2] + 1:
            worksheet.write(3, i, 'Total', format_b)
    for i in range(s[3], e[3] + 2):
        worksheet.write(4, i, 'NP', format_o)
        if i == e[3] + 1:
            worksheet.write(3, i, 'Total', format_o)
    worksheet.write(3, e[3] + 2, 'Deals', format_g)
    worksheet.write(4, e[3] + 2, 'Not in', format_g)
    worksheet.write(5, e[3] + 2, 'Flight', format_g)

    # Group Columns
    worksheet.set_column('D:F', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('H:I', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('M:O', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('AW:AZ', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('BI:DJ', None, None, {'level': 1, 'hidden': 1})

    
    #worksheet.set_column(s_letter[1] + ':' + xlsxwriter.utility.xl_col_to_name(e[1] + 1), None, None, {'level': 1})

    # Autofilter
    worksheet.autofilter('A7:' + xlsxwriter.utility.xl_col_to_name(e[-1] + 2) + str(count_row+7))
    

    # Get the Sum
    for col in range(s[0] - 4, e[3] + 3):
        col = xlsxwriter.utility.xl_col_to_name(col)
        worksheet.write_formula(col + str(count_row + 8),
                                '{=subtotal(9, ' + col + '8:' + col + str(count_row + 6) + ')}')

    # Conditional format for date
    # Add a format. Light red fill with dark red text.
    format1 = workbook.add_format({'bg_color': '#FFC7CE',
                                   'font_color': '#9C0006', 
                                   'font_name': 'Arial'})
    # Add a format. Green fill with dark green text.
    format2 = workbook.add_format({'bg_color': '#C6EFCE',
                                   'font_color': '#006100', 
                                   'font_name': 'Arial'})
    format3 = workbook.add_format({'bg_color': 'white', 
                                   'font_name': 'Arial'})

    worksheet.conditional_format(s_letter[3] + '6:' + e_letter[3] + '6' , {'type': 'cell',
                                                                          'criteria': '<',
                                                                          'value': 'Datevalue($D$1)',
                                                                          'format': format1})
    worksheet.conditional_format(s_letter[3] + '6:' + e_letter[3] + '6', {'type': 'cell',
                                                                          'criteria': '>=',
                                                                          'value': 'Datevalue($D$1)',
                                                                          'format': format2})
    worksheet.conditional_format(s_letter[4] + '6:' + e_letter[4] + '6', {'type': 'cell',
                                                                          'criteria': '<',
                                                                          'value': 'Datevalue($D$1)',
                                                                          'format': format1})
    worksheet.conditional_format(s_letter[4] + '6:' + e_letter[4] + '6', {'type': 'cell',
                                                                          'criteria': '>=',
                                                                          'value': 'Datevalue($D$1)',
                                                                          'format': format2})

    # Column Width
    worksheet.set_column('C' + ':' + e_letter[0], 15)
    worksheet.set_column(0, 2, 2) 
    worksheet.set_column(s_letter[3]+ ':' + e_letter[4], 10) 


    # Freeze the top rows and left columns
    worksheet.freeze_panes(7, 12)

    
    # column Format
    fmt1 = workbook.add_format({'num_format': '0.00', 'font_name': 'Arial'})
    fmt2 = workbook.add_format({'num_format': '#,##0', 'font_name': 'Arial'})
    fmt3 = workbook.add_format({'num_format': '0%', 'font_name': 'Arial'})
    fmt4 = workbook.add_format({'num_format': '0.0', 'font_name': 'Arial'})
    
    worksheet.set_column('O:S', None, fmt2)
    worksheet.set_column('T:T', None, fmt1)
    worksheet.set_column('U:Y', None, fmt2)
    worksheet.set_column('Z:Z', None, fmt1)
    worksheet.set_column('AA:AE', None, fmt2)
    worksheet.set_column('AF:AF', None, fmt1)
    worksheet.set_column('AG:AK', None, fmt2)
    worksheet.set_column('AL:AL', None, fmt1)
    worksheet.set_column('AM:AO', None, fmt2)
    worksheet.set_column('AP:AP', None, fmt3)    
    worksheet.set_column('AQ:AR', None, fmt2)
    worksheet.set_column('AS:AS', None, fmt1)
    worksheet.set_column('AT:AT', None, fmt2)
    worksheet.set_column('AU:AV', None, fmt3)
    worksheet.set_column('AW:AZ', None, fmt2, {'level': 1, 'hidden': 1})
    worksheet.set_column('BA:BB', None, fmt3)
    worksheet.set_column('BC:BD', None, fmt2)
    worksheet.set_column('BE:BG', None, fmt4)
    worksheet.set_column('BH:FL', None, fmt4)
    
    
    writer.save()
    return s, s_letter, e_letter



def format_take_back(raw, new):
    writer = pd.ExcelWriter(DIR_OUTPUT+str(datetime.now().strftime("%Y-%m-%d"))+' ADUs to delete.xlsx', engine='xlsxwriter')#, datetime_format='%m/%d/%Y')
    workbook = writer.book

    # Set Font 
    workbook.formats[0].set_font_name('Arial')
   
    
    count_row = raw[1].shape[0] + 1  # gives number of row count
    count_col = raw[1].shape[1] + 3  # gives number of col count
    raw[1].to_excel(writer, sheet_name='ADUs to delete', startrow=7, startcol=2, header=False, index = False)

    worksheet = writer.sheets['ADUs to delete']
    worksheet.set_zoom(85)

    
    # Clean the headers
    for col_num, value in enumerate(raw[1].columns.values):
        if col_num <= 9:
            worksheet.write(5, col_num + 2, value)
        elif col_num <= 11:
            worksheet.write(5, col_num + 2, ' '.join(value.split()[1:]))
        elif col_num <= 35:
            worksheet.write(5, col_num + 2, ' '.join(value.split()[2:]))
        else:
            worksheet.write(5, col_num + 2, ' '.join(value.split()[1:]))
            
    date_fmt = workbook.add_format({'num_format':'mm/dd/yyyy',
                                    'font_name': 'Arial'})        
    s = [] # stores the start column of each dataframe
    e = [] # stores the end column of each dataframe
    for i in range(2, len(raw)):
        raw[i].iloc[:, 1:].to_excel(writer, sheet_name='ADUs to delete', startrow=7, startcol=count_col, index=False,
                                    header=False)
        if i < 4:
            for col_num, value in enumerate(raw[i].columns.values[1:]):
                worksheet.write(5, count_col + col_num, value)
        else:
            for col_num, value in enumerate(raw[i].columns.values[1:]):
                value = datetime.strptime(value, '%m/%d/%Y')
                worksheet.write_datetime(5, count_col + col_num, value, date_fmt)

        s.append(count_col)
        for r in range(count_row):
            for c in range(count_col, count_col):
                worksheet.write_blank(r, c, None)

        count_col += raw[i].shape[1]
        e.append(count_col - 2)
    s_letter = ['B'] #start column letter of each dataframe
    e_letter = ['L'] #end column letter of each dataframe
    for i in range(len(s)):
        s_letter.append(xlsxwriter.utility.xl_col_to_name(s[i]))
        e_letter.append(xlsxwriter.utility.xl_col_to_name(e[i]))

    # sum of scheduled spots and ADUs
    for i in range(len(e)):
        col = xlsxwriter.utility.xl_col_to_name(e[i] + 1)
        for r in range(8, count_row + 7):
            worksheet.write_formula(col + str(r),
                                    '{=SUM(' + s_letter[i + 1] + str(r) + ':' + e_letter[i + 1] + str(r) + ')}')

    # Deals not in flight
    col = xlsxwriter.utility.xl_col_to_name(e[-1] + 2)
    Total_P_ADU_col = xlsxwriter.utility.xl_col_to_name(e[2] + 1)
    Total_NP_ADU_col = xlsxwriter.utility.xl_col_to_name(e[3] + 1)
    Total_ADU_col = xlsxwriter.utility.xl_col_to_name(s[0] - 2)
    for r in range(8, count_row + 7):
        worksheet.write_formula(col + str(r), '{=' + Total_ADU_col + str(r) + '-' + Total_P_ADU_col + str(
            r) + '-' + Total_NP_ADU_col + str(r) + '}')

    # Take Back Deal
    c = xlsxwriter.utility.xl_col_to_name(e[-1] + 3)
    for r in range(8, count_row + 7):
        worksheet.write_formula(c + str(r), '{=' + Total_P_ADU_col + str(r) + '+' + Total_NP_ADU_col + str(r)+'}')    


    # Header
    bold = workbook.add_format({'bold': True, 
                                'font_name': 'Arial'})
    worksheet.write(1, 2, 'ION Media', bold)
    worksheet.write(2, 2, 'ADU Trust 3.0', bold)
    bold_blue = workbook.add_format({'bold': True, 'font_color': 'blue', 
                                   'font_name': 'Arial'})
    white = workbook.add_format({'font_color': 'white', 
                                    'font_name': 'Arial'})
    worksheet.write(0, 2, str(datetime.now().strftime("%m/%d/%Y")), bold_blue)
    worksheet.write(0, 3, raw[0], white)


    
    # Add Title & Merge
    format_b = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#99CCFF', 
        'font_name': 'Arial'})
    format_o = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#FFCC99', 
        'font_name': 'Arial'})
    format_y = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#FFFFCC', 
        'font_name': 'Arial'})
    format_g = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#C0C0C0', 
        'font_name': 'Arial'})  # grey

    try:
        worksheet.merge_range(s_letter[3] + '4:' + e_letter[3] + '4', 'Prime - ADU Suggested Flighting', format_b)
        worksheet.merge_range(s_letter[4] + '4:' + e_letter[4] + '4', 'Non Prime - ADU Suggested Flighting', format_o)
        worksheet.merge_range(s_letter[1] + '4:' + e_letter[1] + '4', 'Prime Fligting - Sold Units', format_b)
        worksheet.merge_range(s_letter[2] + '4:' + e_letter[2] + '4', 'Non Prime Fligting - Sold Units', format_o)

    except:
        print('nope')

    # Headers for dataframes
    for i in range(2, 59):
        if i <= 13:
            worksheet.write(3, i, 'SOLD', format_g)
            worksheet.write(4, i, ' ', format_g)
        elif i <= 19:
            worksheet.write(3, i, 'SOLD', format_b)
            worksheet.write(4, i, 'Prime', format_b)
        elif i <= 25:
            worksheet.write(3, i, 'SOLD', format_o)
            worksheet.write(4, i, 'NP', format_o)
        elif i <= 31:
            worksheet.write(3, i, 'ADU', format_b)
            worksheet.write(4, i, 'Prime', format_b)
        elif i <= 37:
            worksheet.write(3, i, 'ADU', format_o)
            worksheet.write(4, i, 'NP', format_o)
        elif i <= 47:
            worksheet.write(3, i, 'Total', format_g)
            worksheet.write(4, i, ' ', format_g)
        else:
            worksheet.write(3, i, ' ', format_g)
            if i != 58:
                if i % 2 == 0:
                    worksheet.write(4, i, 'P', format_g)
                else:
                    worksheet.write(4, i, 'NP', format_g)
            else:
                worksheet.write(4, i, 'Total', format_g)
                
                
    for i in range(s[0], e[0] + 2):
        worksheet.write(4, i, 'P', format_b)
        if i == e[0] + 1:
            worksheet.write(3, i, 'Total', format_b)
    for i in range(s[1], e[1] + 2):
        worksheet.write(4, i, 'NP', format_o)
        if i == e[1] + 1:
            worksheet.write(3, i, 'Total', format_o)
    for i in range(s[2], e[2] + 2):
        worksheet.write(4, i, 'P', format_b)
        if i == e[2] + 1:
            worksheet.write(3, i, 'Total', format_b)
    for i in range(s[3], e[3] + 2):
        worksheet.write(4, i, 'NP', format_o)
        if i == e[3] + 1:
            worksheet.write(3, i, 'Total', format_o)

    worksheet.write(3, e[3] + 2, 'Deals', format_g)
    worksheet.write(4, e[3] + 2, 'Not in', format_g)
    worksheet.write(5, e[3] + 2, 'Flight', format_g)

    worksheet.write(3, e[3] + 3, 'Deals', format_g)
    worksheet.write(4, e[3] + 3, 'Take back', format_g)
    worksheet.write(5, e[3] + 3, '', format_g)


    # Group Columns
    worksheet.set_column('D:F', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('H:I', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('M:O', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('AW:AZ', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('BI:DJ', None, None, {'level': 1, 'hidden': 1})


    # Autofilter
    worksheet.autofilter('A7:' + xlsxwriter.utility.xl_col_to_name(e[-1] + 3) + str(count_row+6))
    
    worksheet.filter_column(xlsxwriter.utility.xl_col_to_name(e[-1] + 2), 'x > 0')
   
    # Get the Sum
    for col in range(s[0] - 4, e[3] + 4):
        col = xlsxwriter.utility.xl_col_to_name(col)
        worksheet.write_formula(col + str(count_row + 8),
                                '{=subtotal(9, ' + col + '8:' + col + str(count_row + 7) + ')}')

    # Conditional format for date
    # Add a format. Light red fill with dark red text.
    format1 = workbook.add_format({'bg_color': '#FFC7CE',
                                   'font_color': '#9C0006', 
                                   'font_name': 'Arial'})
    # Add a format. Green fill with dark green text.
    format2 = workbook.add_format({'bg_color': '#C6EFCE',
                                   'font_color': '#006100', 
                                   'font_name': 'Arial'})
    format3 = workbook.add_format({'bg_color': 'white', 
                                    'font_name': 'Arial'})
        

    worksheet.conditional_format(s_letter[3] + '6:' + e_letter[3] + '6' , {'type': 'cell',
                                                                          'criteria': '<',
                                                                          'value': 'Datevalue($D$1)',
                                                                          'format': format1})
    worksheet.conditional_format(s_letter[3] + '6:' + e_letter[3] + '6', {'type': 'cell',
                                                                          'criteria': '>=',
                                                                          'value': 'Datevalue($D$1)',
                                                                          'format': format2})
    worksheet.conditional_format(s_letter[4] + '6:' + e_letter[4] + '6', {'type': 'cell',
                                                                          'criteria': '<',
                                                                          'value': 'Datevalue($D$1)',
                                                                          'format': format1})
    worksheet.conditional_format(s_letter[4] + '6:' + e_letter[4] + '6', {'type': 'cell',
                                                                          'criteria': '>=',
                                                                          'value': 'Datevalue($D$1)',
                                                                          'format': format2})

    # Column Width
    worksheet.set_column('C' + ':' + e_letter[0], 15)
    worksheet.set_column(0, 2, 2) 
    worksheet.set_column(s_letter[3]+ ':' + e_letter[4], 10) 

    # freeze the top rows and left columns
    worksheet.freeze_panes(7, 12)
    

    # column Format
    fmt1 = workbook.add_format({'num_format': '0.00', 'font_name': 'Arial'})
    fmt2 = workbook.add_format({'num_format': '#,##0', 'font_name': 'Arial'})
    fmt3 = workbook.add_format({'num_format': '0%', 'font_name': 'Arial'})
    fmt4 = workbook.add_format({'num_format': '0.0', 'font_name': 'Arial'})
    
    worksheet.set_column('O:S', None, fmt2)
    worksheet.set_column('T:T', None, fmt1)
    worksheet.set_column('U:Y', None, fmt2)
    worksheet.set_column('Z:Z', None, fmt1)
    worksheet.set_column('AA:AE', None, fmt2)
    worksheet.set_column('AF:AF', None, fmt1)
    worksheet.set_column('AG:AK', None, fmt2)
    worksheet.set_column('AL:AL', None, fmt1)
    worksheet.set_column('AM:AO', None, fmt2)
    worksheet.set_column('AP:AP', None, fmt3)    
    worksheet.set_column('AQ:AR', None, fmt2)
    worksheet.set_column('AS:AS', None, fmt1)
    worksheet.set_column('AT:AT', None, fmt2)
    worksheet.set_column('AU:AV', None, fmt3)
    worksheet.set_column('AW:AZ', None, fmt2, {'level': 1, 'hidden': 1})
    worksheet.set_column('BA:BB', None, fmt3)
    worksheet.set_column('BC:BD', None, fmt2)
    worksheet.set_column('BE:BG', None, fmt4)
    worksheet.set_column('BH:FL', None, fmt4)
    

    writer.save()
    return s, s_letter, e_letter

def format_cur_standing(raw, new):
    writer = pd.ExcelWriter(DIR_OUTPUT+str(datetime.now().strftime("%Y-%m-%d"))+' Deal Delivery.xlsx', engine='xlsxwriter')
    workbook = writer.book
    
    # Set Font
    workbook.formats[0].set_font_name('Arial')

    count_row = raw[1].shape[0] + 1  # gives number of row count
    count_col = raw[1].shape[1] + 3  # gives number of col count
    raw[1].to_excel(writer, sheet_name='Deal Delivery', startrow=7, startcol=2, header=False, index = False)
    worksheet = writer.sheets['Deal Delivery']
    worksheet.set_zoom(85)
    
    # Clean the headers
    for col_num, value in enumerate(raw[1].columns.values):
        if col_num <= 9:
            worksheet.write(5, col_num + 2, value)
        elif col_num <= 11:
            worksheet.write(5, col_num + 2, ' '.join(value.split()[1:]))
        elif col_num <= 35:
            worksheet.write(5, col_num + 2, ' '.join(value.split()[2:]))
        else:
            worksheet.write(5, col_num + 2, ' '.join(value.split()[1:]))

    s = [3]
    e = [58]
    s_letter = ['B'] #start column letter of each dataframe
    e_letter = ['L'] #end column letter of each dataframe
    for i in range(len(s)):
        s_letter.append(xlsxwriter.utility.xl_col_to_name(s[i]))
        e_letter.append(xlsxwriter.utility.xl_col_to_name(e[i]))

    
    # Header
    bold = workbook.add_format({'bold': True, 
                                'font_name': 'Arial'})
    worksheet.write(1, 2, 'ION Media', bold)
    worksheet.write(2, 2, 'ADU Trust 3.0', bold)
    bold_blue = workbook.add_format({'bold': True, 'font_color': 'blue', 
                                   'font_name': 'Arial'})
    white = workbook.add_format({'font_color': 'white', 
                                    'font_name': 'Arial'})
    worksheet.write(0, 2, str(datetime.now().strftime("%m/%d/%Y")), bold_blue)
    worksheet.write(0, 3, raw[0], white)

    
    # Add Title & Merge
    format_b = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#99CCFF', 
        'font_name': 'Arial'})
    format_o = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#FFCC99', 
        'font_name': 'Arial'})
    format_y = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#FFFFCC', 
        'font_name': 'Arial'})
    format_g = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#C0C0C0', 
        'font_name': 'Arial'})  # grey


    # Headers for dataframes
    for i in range(2, 59):
        if i <= 13:
            worksheet.write(3, i, 'SOLD', format_g)
            worksheet.write(4, i, ' ', format_g)
        elif i <= 19:
            worksheet.write(3, i, 'SOLD', format_b)
            worksheet.write(4, i, 'Prime', format_b)
        elif i <= 25:
            worksheet.write(3, i, 'SOLD', format_o)
            worksheet.write(4, i, 'NP', format_o)
        elif i <= 31:
            worksheet.write(3, i, 'ADU', format_b)
            worksheet.write(4, i, 'Prime', format_b)
        elif i <= 37:
            worksheet.write(3, i, 'ADU', format_o)
            worksheet.write(4, i, 'NP', format_o)
        elif i <= 47:
            worksheet.write(3, i, 'Total', format_g)
            worksheet.write(4, i, ' ', format_g)
        else:
            worksheet.write(3, i, ' ', format_g)
            if i != 58:
                if i % 2 == 0:
                    worksheet.write(4, i, 'P', format_g)
                else:
                    worksheet.write(4, i, 'NP', format_g)
            else:
                worksheet.write(4, i, 'Total', format_g)
                
                
    worksheet.write(3, e[-1]+1, 'Over', format_g)
    worksheet.write(4, e[-1]+1, 'Delivered', format_g)
    
    worksheet.write(3, e[-1]+2, 'Linked', format_g)
    worksheet.write(4, e[-1]+2, 'Deal', format_g)

    worksheet.write(3, e[-1]+3, 'DealSize', format_g)
    worksheet.write(4, e[-1]+3, 'Group', format_g)

    worksheet.write(3, e[-1]+4, 'Flight', format_g)
    worksheet.write(4, e[-1]+4, 'Length', format_g)

    worksheet.write(3, e[-1]+5, 'FlightLen', format_g)
    worksheet.write(4, e[-1]+5, 'Group', format_g)   
    
    worksheet.write(3, e[-1]+6, 'ADU', format_g)
    worksheet.write(4, e[-1]+6, 'Flag', format_g)
    
    # Current standing flag
    delv_imp_col = xlsxwriter.utility.xl_col_to_name(42)
    c = xlsxwriter.utility.xl_col_to_name(e[-1]+1)
    for r in range(8, count_row + 7):
        worksheet.write_formula(c + str(r), '{=IF('+delv_imp_col+str(r)+'>0, "Over Delv", "Uder Delv")' +'}')   
        
    # Linked Deals Flag
    deal_num_col = xlsxwriter.utility.xl_col_to_name(9)
    c = xlsxwriter.utility.xl_col_to_name(e[-1]+2)
    for r in range(8, count_row + 7):
        worksheet.write_formula(c + str(r), '{=IF(ISNUMBER(SEARCH(",", '+deal_num_col+str(r)+')), '+'"Linked", "Individual")}')   
    
    
    # Deal Size
    deal_size_col = xlsxwriter.utility.xl_col_to_name(38)
    c = xlsxwriter.utility.xl_col_to_name(e[-1]+3)
    for r in range(8, count_row + 7):
        worksheet.write_formula(c + str(r), '{=IF('+deal_size_col+str(r)+'<=100000, "0-100k", ' +\
                                              'IF('+deal_size_col+str(r)+'<=500000, "100k-500k", ' +\
                                              'IF('+deal_size_col+str(r)+'<=1000000, "500k-1M", ' +\
                                              'IF('+deal_size_col+str(r)+'<=2000000, "1M-2M", ' +\
                                              'IF('+deal_size_col+str(r)+'<=4000000, "2M-4M", "4M+")))))}')   
    
    
    # Flight Length
    start_date_col = xlsxwriter.utility.xl_col_to_name(12)
    end_date_col = xlsxwriter.utility.xl_col_to_name(13)
    c = xlsxwriter.utility.xl_col_to_name(e[-1]+4)
    for r in range(8, count_row + 7):
        worksheet.write_formula(c + str(r), '{=ROUND(('+end_date_col+str(r) +'-'+ start_date_col+str(r)+')/7, 0)}')   
        
        
       
    # Flight Length Group
    flight_len_col = c
    c = xlsxwriter.utility.xl_col_to_name(e[-1]+5)
    for r in range(8, count_row + 7):
        worksheet.write_formula(c + str(r), '{=IF('+flight_len_col+str(r)+'<=4, "4", ' +\
                                              'IF('+flight_len_col+str(r)+'<=13, "13", ' +\
                                              'IF('+flight_len_col+str(r)+'<=26, "26", ' +\
                                              'IF('+flight_len_col+str(r)+'<=52, "52", "52+"))))}')   
    
    # With ADU Flag
    NP_ADU_col = xlsxwriter.utility.xl_col_to_name(36)
    P_ADU_col = xlsxwriter.utility.xl_col_to_name(30)
    c = xlsxwriter.utility.xl_col_to_name(e[-1]+6)
    for r in range(8, count_row + 7):
        worksheet.write_formula(c + str(r), '{=IF(' +P_ADU_col+str(r) + '+' + NP_ADU_col+str(r) + '>0, "Yes", "No")}')   
    
    
    # Group Columns
    worksheet.set_column('D:F', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('H:I', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('M:AL', None, None, {'level': 1, 'hidden': 1})

    # Autofilter
    worksheet.autofilter('A7:' + xlsxwriter.utility.xl_col_to_name(e[-1]+6) + str(count_row+7))
    
    # Column Width
    worksheet.set_column('C' + ':' + e_letter[0], 15)
    worksheet.set_column(0, 2, 2) 
    
    # freeze the top rows and left columns
    worksheet.freeze_panes(7, 12)

    # column Format
    fmt1 = workbook.add_format({'num_format': '0.00', 'font_name': 'Arial'})
    fmt2 = workbook.add_format({'num_format': '#,##0', 'font_name': 'Arial'})
    fmt3 = workbook.add_format({'num_format': '0%', 'font_name': 'Arial'})
    fmt4 = workbook.add_format({'num_format': '0.0', 'font_name': 'Arial'})
    
    worksheet.set_column('O:S', None, fmt2, {'level': 1, 'hidden': 1})
    worksheet.set_column('T:T', None, fmt1, {'level': 1, 'hidden': 1})
    worksheet.set_column('U:Y', None, fmt2, {'level': 1, 'hidden': 1})
    worksheet.set_column('Z:Z', None, fmt1, {'level': 1, 'hidden': 1})
    worksheet.set_column('AA:AE', None, fmt2, {'level': 1, 'hidden': 1})
    worksheet.set_column('AF:AF', None, fmt1, {'level': 1, 'hidden': 1})
    worksheet.set_column('AG:AK', None, fmt2, {'level': 1, 'hidden': 1})
    worksheet.set_column('AL:AL', None, fmt1, {'level': 1, 'hidden': 1})
    worksheet.set_column('AM:AO', None, fmt2)
    worksheet.set_column('AP:AP', None, fmt3)    
    worksheet.set_column('AQ:AR', None, fmt2)
    worksheet.set_column('AS:AS', None, fmt1)
    worksheet.set_column('AT:AT', None, fmt2)
    worksheet.set_column('AU:AV', None, fmt3)
    worksheet.set_column('AW:AZ', None, fmt2)
    worksheet.set_column('BA:BB', None, fmt3)
    worksheet.set_column('BC:BD', None, fmt2)
    worksheet.set_column('BE:BG', None, fmt4)
    worksheet.set_column('BH:BM', None, fmt4, {'level': 1, 'hidden': 1})
    
    
    # Add pivot table name
    format_w = workbook.add_format({'font_color': 'white', 'font_name': 'Arial'})
    
    for c in range(2, 65):
        col = xlsxwriter.utility.xl_col_to_name(c)
        worksheet.write_formula(col + '7', '{=CONCATENATE(' + col + '4, '+col + '5, '+col+'6)}', format_w)      
    
    writer.save()
    return s, s_letter, e_letter


def format_forecast_actual(all_ratings, four_q):
    writer = pd.ExcelWriter(DIR_OUTPUT+str(datetime.now().strftime("%Y-%m-%d"))+' Ratings Summary.xlsx', engine='xlsxwriter')
    workbook = writer.book
    
    # Set Font
    workbook.formats[0].set_font_name('Arial')
    
    count_row = all_ratings[1].shape[0] + 1  # gives number of row count
    count_col = all_ratings[1].shape[1] + 3  # gives number of col count
    all_ratings[0].to_excel(writer, sheet_name='Ratings Summary', startrow=5, startcol=2, header=True, index = False)

    worksheet = writer.sheets['Ratings Summary']
    worksheet.set_zoom(85)


    s = [2] # stores the start column of each dataframe
    e = [4] # stores the end column of each dataframe
    for i in range(1, len(all_ratings)):
        all_ratings[i].to_excel(writer, sheet_name='Ratings Summary', startrow=5, startcol=count_col, header=True, index = False)
        s.append(count_col)
        for r in range(count_row):
            for c in range(count_col, count_col):
                worksheet.write_blank(r, c, None)

        count_col += all_ratings[i].shape[1]+1
        e.append(count_col - 2)
    s_letter = [] #start column letter of each dataframe
    e_letter = [] #end column letter of each dataframe
    for i in range(len(s)):
        s_letter.append(xlsxwriter.utility.xl_col_to_name(s[i]))
        e_letter.append(xlsxwriter.utility.xl_col_to_name(e[i]))

    # Forecast vs Actual
    
    forecast_prev_p = xlsxwriter.utility.xl_col_to_name(s[0]+1)
    forecast_prev_np = xlsxwriter.utility.xl_col_to_name(s[0]+2)
    forecast_cur_p = xlsxwriter.utility.xl_col_to_name(s[1]+1)
    forecast_cur_np = xlsxwriter.utility.xl_col_to_name(s[1]+2)
    actual_prev_p = xlsxwriter.utility.xl_col_to_name(s[3]+1)
    actual_prev_np = xlsxwriter.utility.xl_col_to_name(s[3]+2)
    actual_cur_p = xlsxwriter.utility.xl_col_to_name(s[4]+1)
    actual_cur_np = xlsxwriter.utility.xl_col_to_name(s[4]+2)    
    
    
    fore_act_prev_p = xlsxwriter.utility.xl_col_to_name(e[-1]+2)
    for r in range(7, count_row+6):
        worksheet.write_formula(fore_act_prev_p + str(r),'{=' + actual_prev_p + str(r) + '/' + forecast_prev_p + str(r) + '-1' +'}')   
          
    fore_act_prev_np = xlsxwriter.utility.xl_col_to_name(e[-1]+3)
    for r in range(7, count_row+6):
        worksheet.write_formula(fore_act_prev_np + str(r),'{=' + actual_prev_np + str(r) +'/' + forecast_prev_np + str(r) + '-1'+ '}')   

    fore_act_prev_p_delta = xlsxwriter.utility.xl_col_to_name(e[-1]+4)
    for r in range(7, count_row+6):
        worksheet.write_formula(fore_act_prev_p_delta + str(r),'{=' + actual_prev_p + str(r) + '-' + forecast_prev_p + str(r) +'}')   
          
    fore_act_prev_np_delta = xlsxwriter.utility.xl_col_to_name(e[-1]+5)
    for r in range(7, count_row+6):
        worksheet.write_formula(fore_act_prev_np_delta + str(r),'{=' + actual_prev_np + str(r) +'-' + forecast_prev_np + str(r) + '}') 
        
        
        
    fore_act_cur_p = xlsxwriter.utility.xl_col_to_name(e[-1]+7)
    for r in range(7, count_row+6):
        worksheet.write_formula(fore_act_cur_p + str(r),'{=' + actual_cur_p + str(r) +'/' + forecast_cur_p +str(r) + '-1'+ '}')   
          
    fore_act_cur_np = xlsxwriter.utility.xl_col_to_name(e[-1]+8)
    for r in range(7, count_row+6):
        worksheet.write_formula(fore_act_cur_np + str(r),'{=' + actual_cur_np + str(r) + '/' + forecast_cur_np + str(r) + '-1'+ '}')   

    fore_act_cur_p_delta = xlsxwriter.utility.xl_col_to_name(e[-1]+9)
    for r in range(7, count_row+6):
        worksheet.write_formula(fore_act_cur_p_delta + str(r),'{=' + actual_cur_p + str(r) +'-' + forecast_cur_p +str(r) + '}')   
          
    fore_act_cur_np_delta = xlsxwriter.utility.xl_col_to_name(e[-1]+10)
    for r in range(7, count_row+6):
        worksheet.write_formula(fore_act_cur_np_delta + str(r),'{=' + actual_cur_np + str(r) + '-' + forecast_cur_np + str(r) + '}')   
    
    
    new_col_s_letter = [fore_act_prev_p, fore_act_cur_p]
    new_col_e_letter = [fore_act_prev_np_delta, fore_act_cur_np_delta]
    
    
    prev_q = 'Q' + str(four_q[0][0]) + ' ' + str(four_q[0][1])
    cur_q = 'Q' + str(four_q[1][0]) + ' ' + str(four_q[1][1])
    next_q = 'Q' + str(four_q[2][0]) + ' ' + str(four_q[2][1])

    # Add Title & Merge
    format_b = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#99CCFF', 
        'font_name': 'Arial'})
    format_o = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#FFCC99',
        'font_name': 'Arial'})
    format_y = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#FFFFCC',
        'font_name': 'Arial'})
    format_g = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#C0C0C0',
        'font_name': 'Arial'})  # grey

    try:
        worksheet.merge_range(s_letter[0] + '4:' + e_letter[0] + '4', prev_q + ' Forecast', format_o)
        worksheet.merge_range(s_letter[1] + '4:' + e_letter[1] + '4', cur_q + ' Forecast', format_o)
        worksheet.merge_range(s_letter[2] + '4:' + e_letter[2] + '4', next_q + ' Forecast', format_o)
        worksheet.merge_range(s_letter[3] + '4:' + e_letter[3] + '4', prev_q + ' Actual', format_b)
        worksheet.merge_range(s_letter[4] + '4:' + e_letter[4] + '4', cur_q + ' Actual', format_b)
        
        worksheet.merge_range(new_col_s_letter[0] + '4:' + new_col_e_letter[0] + '4', prev_q, format_g)
        worksheet.merge_range(new_col_s_letter[0] + '5:' + new_col_e_letter[0] + '5', 'Actual vs. Forecast', format_g)

        worksheet.merge_range(new_col_s_letter[1] + '4:' + new_col_e_letter[1] + '4', cur_q , format_g)
        worksheet.merge_range(new_col_s_letter[1] + '5:' + new_col_e_letter[1] + '5', 'Actual vs. Forecast', format_g)
 
    except:
        print('nope')
    
    # Formatting Titles
    for i in range(s[0], e[-1]+11):
        if i <= e[2]:
            if i%4!=1:
                worksheet.write(4, i, '', format_o)
            
            if i%4 == 2:
                worksheet.write(5, i, 'Demo', format_o)
            elif i%4 == 3:
                worksheet.write(5, i, 'P Imp', format_o)
            elif i%4 == 0:
                worksheet.write(5, i, 'NP Imp', format_o)
                
        elif i<=e[-1]:
            if i%4!=1:
                worksheet.write(4, i, '', format_b)
            
            if i%4 == 2:
                worksheet.write(5, i, 'Demo', format_b)
            elif i%4 == 3:
                worksheet.write(5, i, 'P Imp', format_b)
            elif i%4 == 0:
                worksheet.write(5, i, 'NP Imp', format_b)
        else: 
            if i%5 == 2:
                worksheet.write(5, i, 'P Imp', format_g)
            elif i%5 == 3:
                worksheet.write(5, i, 'NP Imp', format_g)
            elif i%5 == 4:
                worksheet.write(5, i, 'P Imp Delta', format_g)
            elif i%5 == 0:
                worksheet.write(5, i, 'NP Imp Delta', format_g)

                
    # Format number
    fmt1 = workbook.add_format({'num_format': '0%', 'font_name': 'Arial'})
    worksheet.set_column(new_col_s_letter[0] + ':' + new_col_e_letter[-1], None, fmt1)
    
    fmt2 = workbook.add_format({'num_format': '#,##0', 'font_name': 'Arial'})
    worksheet.set_column(s_letter[0] + ':' + e_letter[-1], None, fmt2)
    worksheet.set_column(fore_act_prev_p_delta + ':' + fore_act_prev_np_delta, None, fmt2)
    worksheet.set_column(fore_act_cur_p_delta + ':' + fore_act_cur_np_delta, None, fmt2)

    worksheet.set_column(0, 2, 2)

    writer.save()
    return s, s_letter, e_letter


def format_ADU_notes(raw):

    notes = pd.read_excel(DIR_INPUT + 'ADU Notes.xlsx', sheets = 'Sheet1')
    
    notes = pd.merge(raw[1], notes, how='left')
    notes = notes.sort_values('Guarantee ID')
    notes = notes[['Notes']]

    writer = pd.ExcelWriter(DIR_OUTPUT+str(datetime.now().strftime("%Y-%m-%d"))+' ADU Notes.xlsx', engine='xlsxwriter')
    workbook = writer.book

    # Set Font 
    workbook.formats[0].set_font_name('Arial')

    count_row = raw[1].shape[0] + 1  # gives number of row count
    count_col = raw[1].shape[1] + 3  # gives number of col count
    raw[1].to_excel(writer, sheet_name='ADU Notes', startrow=7, startcol=2, header=False, index = False)
    
    worksheet = writer.sheets['ADU Notes']
    worksheet.set_zoom(85)

    # Clean the headers
    for col_num, value in enumerate(raw[1].columns.values):
        if col_num <= 9:
            worksheet.write(5, col_num + 2, value)
        elif col_num <= 11:
            worksheet.write(5, col_num + 2, ' '.join(value.split()[1:]))
        elif col_num <= 35:
            worksheet.write(5, col_num + 2, ' '.join(value.split()[2:]))
        else:
            worksheet.write(5, col_num + 2, ' '.join(value.split()[1:]))

    s = [] # stores the start column of each dataframe
    e = [] # stores the end column of each dataframe
    date_fmt = workbook.add_format({'num_format':'mm/dd/yyyy',
                                    'font_name': 'Arial'})     
    for i in range(2, len(raw)):
        raw[i].iloc[:, 1:].to_excel(writer, sheet_name='ADU Notes', startrow=7, startcol=count_col, index=False,
                                    header=False)
        if i < 4:
            for col_num, value in enumerate(raw[i].columns.values[1:]):
                worksheet.write(5, count_col + col_num, value)
        else:
            for col_num, value in enumerate(raw[i].columns.values[1:]):
                value = datetime.strptime(value, '%m/%d/%Y')
                worksheet.write_datetime(5, count_col + col_num, value, date_fmt)
        s.append(count_col)
        for r in range(count_row):
            for c in range(count_col, count_col):
                worksheet.write_blank(r, c, None)

        count_col += raw[i].shape[1]
        e.append(count_col - 2)
    s_letter = ['B'] #start column letter of each dataframe
    e_letter = ['L'] #end column letter of each dataframe
    for i in range(len(s)):
        s_letter.append(xlsxwriter.utility.xl_col_to_name(s[i]))
        e_letter.append(xlsxwriter.utility.xl_col_to_name(e[i]))

    # sum of scheduled spots and ADUs
    for i in range(len(e)):
        col = xlsxwriter.utility.xl_col_to_name(e[i] + 1)
        for r in range(8, count_row + 7):
            worksheet.write_formula(col + str(r),
                                    '{=SUM(' + s_letter[i + 1] + str(r) + ':' + e_letter[i + 1] + str(r) + ')}')
   
    # Deals not in flight
    col = xlsxwriter.utility.xl_col_to_name(e[-1] + 2)
    Total_P_ADU_col = xlsxwriter.utility.xl_col_to_name(e[2] + 1)
    Total_NP_ADU_col = xlsxwriter.utility.xl_col_to_name(e[3] + 1)
    Total_ADU_col = xlsxwriter.utility.xl_col_to_name(s[0] - 2)
    for r in range(8, count_row + 7):
        worksheet.write_formula(col + str(r), '{=' + Total_ADU_col + str(r) + '-' + Total_P_ADU_col + str(
            r) + '-' + Total_NP_ADU_col + str(r) + '}')

    # ADU Notes
    notes.to_excel(writer, sheet_name='ADU Notes', startrow=7, startcol=e[-1]+3, header=False, index = False)

        
        
    # Header
    bold = workbook.add_format({'bold': True, 
                                'font_name': 'Arial'})
    worksheet.write(1, 2, 'ION Media', bold)
    worksheet.write(2, 2, 'ADU Trust 3.0', bold)
    bold_blue = workbook.add_format({'bold': True, 'font_color': 'blue', 
                                   'font_name': 'Arial'})
    white = workbook.add_format({'font_color': 'white', 
                                    'font_name': 'Arial'})
    worksheet.write(0, 2, str(datetime.now().strftime("%m/%d/%Y")), bold_blue)
    worksheet.write(0, 3, raw[0], white)

    
    # Add Title & Merge
    format_b = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#99CCFF', 
        'font_name': 'Arial'})
    format_o = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#FFCC99',
        'font_name': 'Arial'})
    format_y = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#FFFFCC',
        'font_name': 'Arial'})
    format_g = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#C0C0C0',
        'font_name': 'Arial'})  # grey

    try:
        #worksheet.merge_range('C4:I4', 'DEAL', format_g)
        #worksheet.merge_range('C5:I5', ' ', format_g)
        worksheet.merge_range(s_letter[3] + '4:' + e_letter[3] + '4', 'Prime - ADU Suggested Flighting', format_b)
        worksheet.merge_range(s_letter[4] + '4:' + e_letter[4] + '4', 'Non Prime - ADU Suggested Flighting', format_o)
        worksheet.merge_range(s_letter[1] + '4:' + e_letter[1] + '4', 'Prime Fligting - Sold Units', format_b)
        worksheet.merge_range(s_letter[2] + '4:' + e_letter[2] + '4', 'Non Prime Fligting - Sold Units', format_o)

    except:
        print('nope')


    # Headers for dataframes
    for i in range(2, 59):
        if i <= 13:
            worksheet.write(3, i, 'SOLD', format_g)
            worksheet.write(4, i, ' ', format_g)
        elif i <= 19:
            worksheet.write(3, i, 'SOLD', format_b)
            worksheet.write(4, i, 'Prime', format_b)
        elif i <= 25:
            worksheet.write(3, i, 'SOLD', format_o)
            worksheet.write(4, i, 'NP', format_o)
        elif i <= 31:
            worksheet.write(3, i, 'ADU', format_b)
            worksheet.write(4, i, 'Prime', format_b)
        elif i <= 37:
            worksheet.write(3, i, 'ADU', format_o)
            worksheet.write(4, i, 'NP', format_o)
        elif i <= 47:
            worksheet.write(3, i, 'Total', format_g)
            worksheet.write(4, i, ' ', format_g)
        else:
            worksheet.write(3, i, ' ', format_g)
            if i != 58:
                if i % 2 == 0:
                    worksheet.write(4, i, 'P', format_g)
                else:
                    worksheet.write(4, i, 'NP', format_g)
            else:
                worksheet.write(4, i, 'Total', format_g)

    for i in range(s[0], e[0] + 2):
        worksheet.write(4, i, 'P', format_b)
        if i == e[0] + 1:
            worksheet.write(3, i, 'Total', format_b)
    for i in range(s[1], e[1] + 2):
        worksheet.write(4, i, 'NP', format_o)
        if i == e[1] + 1:
            worksheet.write(3, i, 'Total', format_o)
    for i in range(s[2], e[2] + 2):
        worksheet.write(4, i, 'P', format_b)
        if i == e[2] + 1:
            worksheet.write(3, i, 'Total', format_b)
    for i in range(s[3], e[3] + 2):
        worksheet.write(4, i, 'NP', format_o)
        if i == e[3] + 1:
            worksheet.write(3, i, 'Total', format_o)
    worksheet.write(3, e[3] + 2, 'Deals', format_g)
    worksheet.write(4, e[3] + 2, 'Not in', format_g)
    worksheet.write(5, e[3] + 2, 'Flight', format_g)

    worksheet.write(3, e[3] + 3, 'ADU', format_g)
    worksheet.write(4, e[3] + 3, 'Notes', format_g)    
    
    # Group Columns
    worksheet.set_column('D:F', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('H:I', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('M:O', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('AW:AZ', None, None, {'level': 1, 'hidden': 1})
    worksheet.set_column('BI:DJ', None, None, {'level': 1, 'hidden': 1})

    
    #worksheet.set_column(s_letter[1] + ':' + xlsxwriter.utility.xl_col_to_name(e[1] + 1), None, None, {'level': 1})

    # Autofilter
    worksheet.autofilter('A7:' + xlsxwriter.utility.xl_col_to_name(e[-1] + 2) + str(count_row+8))
    

    # Get the Sum
    for col in range(s[0] - 4, e[3] + 3):
        col = xlsxwriter.utility.xl_col_to_name(col)
        worksheet.write_formula(col + str(count_row + 8),
                                '{=subtotal(9, ' + col + '8:' + col + str(count_row + 6) + ')}')

    # Conditional format for date
    # Add a format. Light red fill with dark red text.
    format1 = workbook.add_format({'bg_color': '#FFC7CE',
                                   'font_color': '#9C0006', 
                                   'font_name': 'Arial'})
    # Add a format. Green fill with dark green text.
    format2 = workbook.add_format({'bg_color': '#C6EFCE',
                                   'font_color': '#006100', 
                                   'font_name': 'Arial'})
    format3 = workbook.add_format({'bg_color': 'white', 
                                   'font_name': 'Arial'})

    worksheet.conditional_format(s_letter[3] + '6:' + e_letter[3] + '6' , {'type': 'cell',
                                                                          'criteria': '<',
                                                                          'value': 'Datevalue($D$1)',
                                                                          'format': format1})
    worksheet.conditional_format(s_letter[3] + '6:' + e_letter[3] + '6', {'type': 'cell',
                                                                          'criteria': '>=',
                                                                          'value': 'Datevalue($D$1)',
                                                                          'format': format2})
    worksheet.conditional_format(s_letter[4] + '6:' + e_letter[4] + '6', {'type': 'cell',
                                                                          'criteria': '<',
                                                                          'value': 'Datevalue($D$1)',
                                                                          'format': format1})
    worksheet.conditional_format(s_letter[4] + '6:' + e_letter[4] + '6', {'type': 'cell',
                                                                          'criteria': '>=',
                                                                          'value': 'Datevalue($D$1)',
                                                                          'format': format2})

    # Column Width
    worksheet.set_column('C' + ':' + e_letter[0], 15)
    worksheet.set_column(0, 2, 2) 
    worksheet.set_column(s_letter[3]+ ':' + e_letter[4], 10) 
    worksheet.set_column('DN:DN', 40)


    # Freeze the top rows and left columns
    worksheet.freeze_panes(7, 12)

    
    # column Format
    fmt1 = workbook.add_format({'num_format': '0.00', 'font_name': 'Arial'})
    fmt2 = workbook.add_format({'num_format': '#,##0', 'font_name': 'Arial'})
    fmt3 = workbook.add_format({'num_format': '0%', 'font_name': 'Arial'})
    fmt4 = workbook.add_format({'num_format': '0.0', 'font_name': 'Arial'})
    
    worksheet.set_column('O:S', None, fmt2)
    worksheet.set_column('T:T', None, fmt1)
    worksheet.set_column('U:Y', None, fmt2)
    worksheet.set_column('Z:Z', None, fmt1)
    worksheet.set_column('AA:AE', None, fmt2)
    worksheet.set_column('AF:AF', None, fmt1)
    worksheet.set_column('AG:AK', None, fmt2)
    worksheet.set_column('AL:AL', None, fmt1)
    worksheet.set_column('AM:AO', None, fmt2)
    worksheet.set_column('AP:AP', None, fmt3)    
    worksheet.set_column('AQ:AR', None, fmt2)
    worksheet.set_column('AS:AS', None, fmt1)
    worksheet.set_column('AT:AT', None, fmt2)
    worksheet.set_column('AU:AV', None, fmt3)
    worksheet.set_column('AW:AZ', None, fmt2, {'level': 1, 'hidden': 1})
    worksheet.set_column('BA:BB', None, fmt3)
    worksheet.set_column('BC:BD', None, fmt2)
    worksheet.set_column('BE:BG', None, fmt4)
    worksheet.set_column('BH:FL', None, fmt4)
    
    
    writer.save()
    return s, s_letter, e_letter


def format_liability_qtr_report(date, report_df):
    writer = pd.ExcelWriter(DIR_OUTPUT+str(datetime.now().strftime("%Y-%m-%d"))+' Liability Report.xlsx', engine='xlsxwriter')#, datetime_format='%m/%d/%Y')
    workbook = writer.book

    # Set Font 
    workbook.formats[0].set_font_name('Arial')
   
    report_df.to_excel(writer, sheet_name='Liability Report', startrow=7, startcol=2, header=False, index = False)

    worksheet = writer.sheets['Liability Report']
    worksheet.set_zoom(85)

    # Header
    bold = workbook.add_format({'bold': True, 
                                'font_name': 'Arial'})
    worksheet.write(1, 2, 'ION Media', bold)
    worksheet.write(2, 2, 'ADU Trust 3.0', bold)
    bold_blue = workbook.add_format({'bold': True, 'font_color': 'blue', 
                                   'font_name': 'Arial'})
    white = workbook.add_format({'font_color': 'white', 
                                    'font_name': 'Arial'})
    worksheet.write(0, 2, str(datetime.now().strftime("%m/%d/%Y")), bold_blue)
    worksheet.write(0, 3, date, white)

    s = 2
    e = 2 + len(report_df.columns)

    # Add Title & Merge
    format_b = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#99CCFF', 
        'font_name': 'Arial'})
    format_o = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#FFCC99', 
        'font_name': 'Arial'})
    format_y = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#FFFFCC', 
        'font_name': 'Arial'})
    format_g = workbook.add_format({
        'bold': 1,
        'align': 'left',
        'valign': 'vcenter',
        'fg_color': '#C0C0C0', 
        'font_name': 'Arial'})  # grey
    

    # Headers for dataframes
    for i in range(s, e):
        worksheet.write(6, i, report_df.columns[i-2], format_g)
          
    a = xlsxwriter.utility.xl_col_to_name(e-4)
    worksheet.set_column('J:'+a, None, None, {'level': 1, 'hidden': 1})

    # Autofilter
    worksheet.autofilter('A7:' + xlsxwriter.utility.xl_col_to_name(e-1) + str(len(report_df)+7))
       
    # Column Width
    worksheet.set_column('C' + ':' + xlsxwriter.utility.xl_col_to_name(e), 15)
    worksheet.set_column(0, 2, 2) 

    # freeze the top rows and left columns
    worksheet.freeze_panes(7, 9)

    # column Format
    fmt1 = workbook.add_format({'num_format': '0%', 'font_name': 'Arial'})
    worksheet.set_column('J:'+a, None, fmt1, {'level': 1, 'hidden': 1})
    worksheet.set_column(a + ':' + xlsxwriter.utility.xl_col_to_name(e), None, fmt1)

    writer.save()
    return 



def new_data(raw, quarters):
    general = dict()
    P_ADU_dict = raw[4].to_dict('index')
    NP_ADU_dict = raw[5].to_dict('index')

    for k, v in P_ADU_dict.items():
        gid=v['Guarantee ID']
        if gid not in general:
            general[gid] = {'Year': [], 'Quarter': [], 'Year + Quarter': [], 'Week Start Date': [],
                                        'Week End Date': [],  'Selling Title': [], \
                                        'Days And Times': [], 'ADU Ind': [], 'Booked Dollars': [],
                                        'Primary Demo Equiv Deal Imp': [], \
                                        #'Primary Demo Equiv Post Imp - IE 1': [],
                                        'Primary Demo Non-ADU Equiv Deal Imp': [], \
                                        #'Primary Demo Equiv Ratecard Imp': [], 
                                        'Primary Demo Deal CPM': [], \
                                        'Equiv Units': []}
        for key, value in v.items():
            if key != 'Guarantee ID' and value > 0:
                # filling all the information
                general[gid]['Week Start Date'].append(key)
                general[gid]['Equiv Units'].append(value)

                y = key.split('/')[2]
                general[gid]['Year'].append(y)

                mon = pd.date_range(key, periods=1, freq='W-MON').strftime('%m/%d/%Y').tolist()[0]
                q = quarters.loc[quarters['start_date'].astype(str) == mon, 'quarter'].iloc[0][1]

                general[gid]['Quarter'].append(q)
                general[gid]['Year + Quarter'].append(y + ' ' + str(q) + 'Q')
                general[gid]['Week End Date'].append(pd.date_range(key, periods=1, freq='W-SUN').strftime('%m/%d/%Y').tolist()[0])
                general[gid]['Selling Title'].append('P')
                general[gid]['Days And Times'].append('')
                general[gid]['ADU Ind'].append('Y')
                general[gid]['Booked Dollars'].append(0)
                general[gid]['Primary Demo Equiv Deal Imp'].append(0)
                #general[gid]['Primary Demo Equiv Post Imp - IE 1'].append(0)
                general[gid]['Primary Demo Non-ADU Equiv Deal Imp'].append(0)
                #general[gid]['Primary Demo Equiv Ratecard Imp'].append(0)
                general[gid]['Primary Demo Deal CPM'].append(0)
                
               
    for k, v in NP_ADU_dict.items():
        gid=v['Guarantee ID']
        if gid not in general:
            general[gid] = {'Year': [], 'Quarter': [], 'Year + Quarter': [], 'Week Start Date': [],
                                        'Week End Date': [],  'Selling Title': [], \
                                        'Days And Times': [], 'ADU Ind': [], 'Booked Dollars': [],
                                        'Primary Demo Equiv Deal Imp': [], \
                                        #'Primary Demo Equiv Post Imp - IE 1': [],
                                        'Primary Demo Non-ADU Equiv Deal Imp': [], \
                                        #'Primary Demo Equiv Ratecard Imp': [], 
                                        'Primary Demo Deal CPM': [], \
                                        'Equiv Units': []}
        for key, value in v.items():
            if key != 'Guarantee ID' and value > 0:
                general[gid]['Week Start Date'].append(key)
                general[gid]['Equiv Units'].append(value)

                y = key.split('/')[2]
                general[gid]['Year'].append(y)

                mon = pd.date_range(key, periods=1, freq='W-MON').strftime('%m/%d/%Y').tolist()[0]
                q = quarters.loc[quarters['start_date'].astype(str) == mon, 'quarter'].iloc[0][1]

                general[gid]['Quarter'].append(q)
                general[gid]['Year + Quarter'].append(y + ' ' + str(q) + 'Q')
                general[gid]['Week End Date'].append(pd.date_range(key, periods=1, freq='W-SUN').strftime('%m/%d/%Y').tolist()[0])
                general[gid]['Selling Title'].append('NP')
                general[gid]['Days And Times'].append('')
                general[gid]['ADU Ind'].append('Y')
                general[gid]['Booked Dollars'].append(0)
                general[gid]['Primary Demo Equiv Deal Imp'].append(0)
                #general[gid]['Primary Demo Equiv Post Imp - IE 1'].append(0)
                general[gid]['Primary Demo Non-ADU Equiv Deal Imp'].append(0)
                #general[gid]['Primary Demo Equiv Ratecard Imp'].append(0)
                general[gid]['Primary Demo Deal CPM'].append(0)
                
    return general

def newdata_to_df(df, general, output):
    df['In System'] = 'Y'
    basics = df[['Guarantee ID', 'Guarantee Name', 'Deal Numbers in Guarantee', 'Marketplace', \
                 'Advertiser', \
                 'AE Name', 'Agency Name (Billing)', 'Deal Name', 'Deal Number', 'Deal Flight Start Date', \
                 'Deal Flight End Date', 'Deal Year', 'Primary Demo']]
    
    # Create a dataframe from general dictionary
    column_names = ['Guarantee ID', 'Year', 'Quarter', 'Year + Quarter', 'Week Start Date', 'Week End Date',
                    'Selling Title', \
                    'Days And Times', 'ADU Ind', 'Booked Dollars', 'Primary Demo Equiv Deal Imp', \
                    #'Primary Demo Equiv Post Imp - IE 1', \
                    'Primary Demo Non-ADU Equiv Deal Imp', \
                    #'Primary Demo Equiv Ratecard Imp', 
                    'Primary Demo Deal CPM', \
                    'Equiv Units']
    rows = []
    for k, v in general.items():
        for i in range(len(v['Year'])):
            row = []
            row.append(k)
            for key in column_names[1:]:
                row.append(v[key][i])
            rows.append(row)

    newdata_df = pd.DataFrame(rows)

    newdata_df.columns = column_names
    newdata_df['In System'] = 'N'
    
    
    # To get the basic information for new data
    combined = pd.merge(newdata_df, basics, how='left', on='Guarantee ID').drop_duplicates(
        subset=['Guarantee ID', 'Week Start Date', 'Week End Date', 'Selling Title'])
    
    # To get impression for new data
    imp_df = output[['Guarantee ID', 'P Forecast Imp', 'NP Forecast Imp']]
    combined = pd.merge(combined, imp_df, on='Guarantee ID')
    
    imp = dict()
    for i, r in combined.iterrows():
        if r['Selling Title'] == 'P':
            imp[(r['Guarantee ID'], 'P', r['Equiv Units'])] = r['P Forecast Imp'] * r['Equiv Units'] * 1000
        else:
            imp[(r['Guarantee ID'], 'NP', r['Equiv Units'])] = r['NP Forecast Imp'] * r['Equiv Units'] * 1000

    ADU_E_D_I = pd.Series(imp).rename_axis(['Guarantee ID', 'Selling Title', 'Equiv Units']).reset_index(
        name='Primary Demo Equiv Post Imp')

    combined = pd.merge(combined, ADU_E_D_I, how='left', on=['Guarantee ID', 'Selling Title', 'Equiv Units'])
    #combined['Primary Demo Equiv Post Imp'] = combined['Primary Demo ADU Equiv Deal Imp']
    
    combined = combined[['Guarantee ID', 'Guarantee Name', 'Deal Numbers in Guarantee', 'Marketplace', \
                         'Advertiser', \
                         'AE Name', 'Agency Name (Billing)', 'Deal Name', 'Deal Number', 'Deal Flight Start Date', \
                         'Deal Flight End Date', 'Deal Year', 'Primary Demo', 'Year', 'Quarter', 'Year + Quarter', 'Week Start Date', \
                         'Week End Date', 'Selling Title', 'Days And Times', 'ADU Ind', 'Booked Dollars', \
                         'Primary Demo Equiv Deal Imp', \
                         #'Primary Demo Equiv Post Imp - IE 1', \
                         'Primary Demo Non-ADU Equiv Deal Imp', \
                         #'Primary Demo Equiv Ratecard Imp', \
                         'Primary Demo Equiv Post Imp', 'Primary Demo Deal CPM', 'Equiv Units', \
                         'In System']]

    total = pd.concat([df, combined], sort=False)

    return total


def liability(new):
    # Sort the dataframe
    new['Week Start Date'] =  pd.to_datetime(new['Week Start Date'])
    new['Week End Date'] =  pd.to_datetime(new['Week End Date'])
    df_sort = new.sort_values(['Guarantee ID','In System', 'Week Start Date', 'Week End Date', 'Booked Dollars'],
                              ascending=[True, False, True, True, False])
    Acc_Deal_Imp = set()
    Acc_Deal_Imp_list = []
    Acc_Delv_Imp = []
    ACC_Effec_Delv_Imp = []
    Effec_Delv_Imp = []
    Owed_Imp = []

    Acc_Deal_value = []
    Effec_Delv_value = []
    Acc_Effec_Delv_value = []
    Owed_value = []
    
    # Compute impressions and values
    for i, r in df_sort.iterrows():

        if r['Guarantee ID'] not in Acc_Deal_Imp:
            Acc_Deal_Imp.add(r['Guarantee ID'])
            a = r['Primary Demo Non-ADU Equiv Deal Imp']
            b = r['Primary Demo Equiv Post Imp']
            c = min(a, b)
            d = c

            A = float(r['Booked Dollars'])

            pool = []            
            Guar = r['Primary Demo Non-ADU Equiv Deal Imp'] * float(r['Primary Demo Deal CPM'])
            pool.append([r['Primary Demo Non-ADU Equiv Deal Imp'], float(r['Primary Demo Deal CPM'])])
            B = d * float(r['Primary Demo Deal CPM'])

            pool[0][0] -= d
            C = B


        else:
            a = Acc_Deal_Imp_list[-1] + r['Primary Demo Non-ADU Equiv Deal Imp']
            b = Acc_Delv_Imp[-1] + r['Primary Demo Equiv Post Imp']
            c = min(a, b)
            d = c - ACC_Effec_Delv_Imp[-1]

            A = Acc_Deal_value[-1] + float(r['Booked Dollars'])

            Guar = r['Primary Demo Non-ADU Equiv Deal Imp'] * float(r['Primary Demo Deal CPM'])
            pool.append([r['Primary Demo Non-ADU Equiv Deal Imp'], float(r['Primary Demo Deal CPM'])])

            B = 0
            imp = d
            while imp > pool[0][0]:
                temp = pool.pop(0)
                B += temp[0] * temp[1]
                imp -= temp[0]
            if imp > 0:
                B += imp * pool[0][1]
                pool[0][0] -= imp

            C = Acc_Effec_Delv_value[-1] + B

        e = r['Primary Demo Non-ADU Equiv Deal Imp'] - d
        D = Guar - B

        Acc_Deal_Imp_list.append(a)
        Acc_Delv_Imp.append(b)
        ACC_Effec_Delv_Imp.append(c)
        Effec_Delv_Imp.append(d)
        Owed_Imp.append(e)

        Acc_Deal_value.append(A)
        Effec_Delv_value.append(B)
        Acc_Effec_Delv_value.append(C)
        Owed_value.append(D)

    df_sort['Acc_Deal_Imp'] = Acc_Deal_Imp_list
    df_sort['Acc_Delv_Imp'] = Acc_Delv_Imp
    df_sort['Acc_Effec_Delv_Imp'] = ACC_Effec_Delv_Imp
    df_sort['Effec_Delv_Imp'] = Effec_Delv_Imp
    df_sort['Owed_Imp'] = Owed_Imp

    df_sort['Acc_Deal_value'] = Acc_Deal_value
    df_sort['Effec_Delv_value'] = Effec_Delv_value
    df_sort['Acc_Effec_Delv_value'] = Acc_Effec_Delv_value
    df_sort['Owed_value'] = Owed_value

    df_sort['Primary Demo Equiv Deal Imp'] = df_sort['Primary Demo Equiv Deal Imp']/1000
    #df_sort['Primary Demo Equiv Post Imp - IE 1'] = df_sort['Primary Demo Equiv Post Imp - IE 1'] / 1000
    #df_sort['Primary Demo ADU Equiv Deal Imp'] = df_sort['Primary Demo ADU Equiv Deal Imp'] / 1000
    df_sort['Primary Demo Non-ADU Equiv Deal Imp'] = df_sort['Primary Demo Non-ADU Equiv Deal Imp'] / 1000
    #df_sort['Primary Demo Equiv Ratecard Imp'] = df_sort['Primary Demo Equiv Ratecard Imp'] / 1000
    df_sort['Primary Demo Equiv Post Imp'] = df_sort['Primary Demo Equiv Post Imp'] / 1000
    df_sort['Acc_Deal_Imp'] = df_sort['Acc_Deal_Imp'] / 1000
    df_sort['Acc_Delv_Imp'] = df_sort['Acc_Delv_Imp'] / 1000
    df_sort['Acc_Effec_Delv_Imp'] = df_sort['Acc_Effec_Delv_Imp'] / 1000
    df_sort['Effec_Delv_Imp'] = df_sort['Effec_Delv_Imp']/1000
    df_sort['Owed_Imp'] = df_sort['Owed_Imp']/1000
    
    df_sort['Effec_Delv_value'] = df_sort['Effec_Delv_value']/1000
    df_sort['Acc_Effec_Delv_value'] = df_sort['Acc_Effec_Delv_value']/1000
    df_sort['Owed_value'] = df_sort['Owed_value']/1000


    return df_sort

def calc_units(liab, raw):
    info = raw[1][['Guarantee ID', 'Total P Mix %', 'P Forecast Imp', 'NP Forecast Imp']]
    liab_update = pd.merge(liab, info, how='left', on='Guarantee ID')
    liab_update['Effective_ADU'] = liab_update['Owed_Imp']*liab_update['Total P Mix %']/liab_update['P Forecast Imp'] \
                                + liab_update['Owed_Imp']*(1-liab_update['Total P Mix %'])/liab_update['NP Forecast Imp']
    return liab_update

def combine_demo(df):
    demos = df['Primary Demo'].unique().tolist()
    startpoint = ['2','6','9','12','15','18','21','25','30','35','40','45', '50','55','65']
    demo_list = ['2-5','6-8','9-11','12-14', '15-17', '18-20', '21-24', '25-29', '30-34', '35-39', '40-44', '45-49', '50-54', '55-64', '65+']
    demo_dic = {'HH':['HHLD']}
    for d in demos:
        if d != 'HH':
            demo_dic[d] = []
            gen = d[0]
            if '+' not in d:
                s, e = d[1:].split('-')
                #print(s,e)
                s_ind = startpoint.index(s)
                e_ind = startpoint.index(str(int(e)+1))
                if gen != 'P':
                    for i in range(s_ind, e_ind):
                        demo_dic[d].append(gen + demo_list[i])
                else:
                    for i in range(s_ind, e_ind):
                        demo_dic[d].append('F' + demo_list[i])
                        demo_dic[d].append('M' + demo_list[i])

            else:
                s = d[1:-1]
                s_ind = startpoint.index(s)
                if gen != 'P':
                    for i in range(s_ind, len(demo_list)):
                        demo_dic[d].append(gen + demo_list[i])
                else:
                    for i in range(s_ind, len(demo_list)):
                        demo_dic[d].append('F' + demo_list[i])
                        demo_dic[d].append('M' + demo_list[i])
    return demo_dic

def get_ratings(df, internal_estimates, cur_q, cur_y):
    estimates_NP = internal_estimates.loc[(internal_estimates['Selling Title'] == 'MSU7A7P1A3A') & (internal_estimates['Quarter'] ==cur_q ) & (internal_estimates['Year'] ==cur_y )]
    estimates_P = internal_estimates.loc[(internal_estimates['Selling Title'] == 'MSU7p1a')& (internal_estimates['Quarter'] ==cur_q) & (internal_estimates['Year'] ==cur_y )]
    demo_dic = combine_demo(df)
    
    demo_ratings_P = dict()
    demo_ratings_NP = dict()

    rating_dic_P = list(estimates_P.to_dict(orient = 'index').values())[0]
    rating_dic_NP = list(estimates_NP.to_dict(orient = 'index').values())[0]

    for k,v in demo_dic.items():
        demo_ratings_P[k] = 0
        demo_ratings_NP[k] = 0

        for d in v:
            demo_ratings_P[k] += rating_dic_P[d]
            demo_ratings_NP[k] += rating_dic_NP[d]

    demo_ratings_P = pd.DataFrame.from_dict(demo_ratings_P, orient='index').reset_index()
    demo_ratings_NP = pd.DataFrame.from_dict(demo_ratings_NP, orient='index').reset_index()
    ratings = pd.merge(demo_ratings_P, demo_ratings_NP, on='index')
    ratings.columns = ['Demo', 'Prime Imp', 'Non Prime Imp']
    return ratings


def copy_rename(old_file_name, new_file_name):
    src_dir= DIR_INPUT
    src_file = os.path.join(src_dir, old_file_name)
    dst_dir= DIR_ARCHIVE
    dst_file = os.path.join(dst_dir, old_file_name)
    new_dst_file_name = os.path.join(dst_dir, new_file_name)

    shutil.copy(src_file,dst_dir)
    os.rename(dst_file, new_dst_file_name)
    os.remove(src_file)
    return


def seperate(raw):
    df = raw[1].copy()
    pos = df[df['Total Imps Owed']>=0]
    gid = pos['Guarantee ID'].tolist()
    df_p = raw[4].copy()
    df_p['sum'] = df_p.select_dtypes(float).sum(1)
    
    df_np = raw[5].copy()
    df_np['sum'] = df_np.select_dtypes(float).sum(1)
    
    P_take_back = df_p[df_p['sum']<0]
    NP_take_back = df_np[df_np['sum']<0]
    
    P_gid = P_take_back['Guarantee ID'].tolist()
    NP_gid = NP_take_back['Guarantee ID'].tolist()
    
    neg_gid = set(P_gid).union(set(NP_gid))
    
    sch = raw[0], pos, \
        raw[2][raw[2]['Guarantee ID'].isin(gid)], \
        raw[3][raw[3]['Guarantee ID'].isin(gid)], \
        raw[4][raw[4]['Guarantee ID'].isin(gid)], \
        raw[5][raw[5]['Guarantee ID'].isin(gid)]

    takeback = raw[0],\
        raw[1][raw[1]['Guarantee ID'].isin(neg_gid)], \
        raw[2][raw[2]['Guarantee ID'].isin(neg_gid)], \
        raw[3][raw[3]['Guarantee ID'].isin(neg_gid)], \
        raw[4][raw[4]['Guarantee ID'].isin(neg_gid)], \
        raw[5][raw[5]['Guarantee ID'].isin(neg_gid)]

    return sch, takeback


def get_report_values(liab, four_q):
  
    quar = sorted(liab['Year + Quarter'].unique())
   
    cur_year_quar = str(four_q[2][1]) + ' ' + four_q[2][0] + 'Q'
    quar = quar[:quar.index(cur_year_quar)+1]
   
        
    table1 = pd.pivot_table(liab[liab['In System']=='Y'], index = 'Year + Quarter', columns = 'ADU Ind', values=['Owed_value', 'Owed_Imp', 'Equiv Units', 'Effective_ADU'], aggfunc=np.sum, fill_value=0, margins = True)
    table2 = pd.pivot_table(liab, index = 'Year + Quarter', columns = ['ADU Ind'], values=['Owed_value', 'Owed_Imp', 'Equiv Units', 'Effective_ADU'], aggfunc=np.sum, fill_value=0, margins = True)

    begin_liab = []
    begin_imp_owed = []
    begin_adu_req = []

    cur_q_liab = []
    cur_q_imp_owed = []
    cur_q_adu_req = []

    cur_q_liab_paid = []
    cur_q_imp_paid = []
    cur_q_adu_given = []
    cur_q_liab_paid_new = []
    cur_q_imp_paid_new = []
    cur_q_adu_given_new = []

    ending_liab = []
    ending_imp_owed = []
    ending_adu_req =[]
    ending_liab_new = []
    ending_imp_owed_new = []
    ending_adu_req_new = []


    table1.reset_index(inplace=True)

    order = table1['Year + Quarter'].tolist()

    owed_v_spots = table1['Owed_value']['N'].tolist()
    owed_v_adu = table1['Owed_value']['Y'].tolist()
    owed_v_total = table1['Owed_value']['All'].tolist()

    owed_imp_spots = table1['Owed_Imp']['N'].tolist()
    owed_imp_adu = table1['Owed_Imp']['Y'].tolist()
    owed_imp_total = table1['Owed_Imp']['All'].tolist()
    
    effective_adu_total = table1['Effective_ADU']['All'].tolist()

    adu_units = table1['Equiv Units']['Y'].tolist()

    owed_v_adu_new = table2['Owed_value']['Y'].tolist()
    owed_imp_adu_new = table2['Owed_Imp']['Y'].tolist()
    adu_units_new = table2['Equiv Units']['Y'].tolist()

    owed_v_total_new = table2['Owed_value']['All'].tolist()
    owed_imp_total_new = table2['Owed_Imp']['All'].tolist()
    
    effective_adu_total_new = table2['Effective_ADU']['All'].tolist()

    i = order.index(quar[0])
    for j in range(i, i+len(quar)):
        begin_liab.append(sum(owed_v_total[:j]))
        begin_imp_owed.append(sum(owed_imp_total[:j]))
        begin_adu_req.append(sum(effective_adu_total[:j]))

        ending_liab.append(sum(owed_v_total[:j+1]))
        ending_imp_owed.append(sum(owed_imp_total[:j+1]))
        ending_adu_req.append(sum(effective_adu_total[:j+1]))
        ending_liab_new.append(sum(owed_v_total_new[:j+1]))
        ending_imp_owed_new.append(sum(owed_imp_total_new[:j+1]))
        ending_adu_req_new.append(sum(effective_adu_total_new[:j+1]))

    for q in quar:
        i = order.index(q)

        cur_q_liab.append(owed_v_spots[i])
        cur_q_imp_owed.append(owed_imp_spots[i])
        
        cur_q_liab_paid.append(-owed_v_adu[i])
        cur_q_imp_paid.append(-owed_imp_adu[i])
        cur_q_adu_given.append(adu_units[i])
        cur_q_liab_paid_new.append(-owed_v_adu_new[i])
        cur_q_imp_paid_new.append(-owed_imp_adu_new[i])
        cur_q_adu_given_new.append(adu_units_new[i])

    for i in range(len(quar)):
        cur_q_adu_req.append(ending_adu_req[i]-begin_adu_req[i]+cur_q_adu_given[i])

        
    return (quar, (begin_liab, begin_imp_owed, begin_adu_req, cur_q_liab, cur_q_imp_owed , cur_q_adu_req, cur_q_liab_paid,\
cur_q_imp_paid, cur_q_adu_given, cur_q_liab_paid_new, cur_q_imp_paid_new, cur_q_adu_given_new,\
ending_liab, ending_imp_owed, ending_adu_req, ending_liab_new, ending_imp_owed_new, ending_adu_req_new))



from openpyxl.utils import get_column_letter

def get_summary(report_values, date_string, quar):

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Summary")
    ws.sheet_view.zoomScale = 85
    
    row_start = ws.max_row + 4
    
    #Header
    ws.cell(1, 2).value = datetime.now().strftime("%m/%d/%Y")
    ws.cell(1, 2).font = Font(bold=True, color=colors.BLUE, name = 'Arial')
    ws.cell(2, 2).value = 'ION Media'
    ws.cell(2, 2).font = Font(bold=True, name = 'Arial')    
    ws.cell(3, 2).value = 'ADU Trust 3.0'
    ws.cell(3, 2).font = Font(name = 'Arial') 

    
    #Set column width
    ws.column_dimensions['A'].width = 8
    
    column = 2
    space = ['C', 'G', 'K', 'R']
    while column < 25:
        i = get_column_letter(column)
        if i not in space:
            ws.column_dimensions[i].width = 12
        else:
            ws.column_dimensions[i].width = 3
        column += 1
    
    # write date and year+quarter  
    ws.cell(row_start, 2).value = datetime.now().strftime("%m/%d/%Y")
    ws.cell(row_start, 2).font = Font(bold=True, color=colors.RED, name = 'Arial')
    ws.cell(row_start, 2).fill = PatternFill("solid", fgColor=colors.YELLOW)
    
    for i in range(2, 2+len(quar)):
        ws.cell(row_start+i, 2).value = quar[i-2]
        ws.cell(row_start+i, 2).font = Font(name = 'Arial')
    
    # headers
    h1 = ['Qtr Begin', 'Qtr Begin', 'Qtr Begin', 'Current Qtr', 'Current Qtr', 'Current Qtr', \
          'Current Qtr', 'Current Qtr', 'Current Qtr', 'Current Qtr', 'Current Qtr', 'Current Qtr', \
          'Qtr End','Qtr End', 'Qtr End', 'Qtr End', 'Qtr End', 'Qtr End']
    h2 = ['Liability', 'Impression Owed', 'ADU required', 'Liability', 'Impression Owed', 'ADU required',\
         'Liaility Paid', 'Impression Paid', 'ADUs Given', 'Liaility Paid(new)','Impression Paid (new sch)', 'ADUs Given(new sch)', \
         'Liability Bal', 'Impression owed', 'ADUs Required','Liaility Bal(new)', 'Impression owed(new sch)', 'ADUs Required(new Sch)']
    
    wb.save(DIR_OUTPUT+'Summary.xlsx')
    
    val_i = 0
    for col_i in range(4, 25):
        if col_i in {7,11,18}:
            pass
        else:
            ws.cell(row_start, col_i).value = h1[val_i]
            ws.cell(row_start, col_i).font = Font(bold=True, underline="single", name = 'Arial')
            ws.cell(row_start+1, col_i).value = h2[val_i]
            ws.cell(row_start+1, col_i).font = Font(bold=True, underline="single", name = 'Arial')
        
            for r in range(row_start+2, row_start+2+len(quar)):
                ws.cell(r, col_i).value = report_values[val_i][r-row_start-2]
                ws.cell(r, col_i).font = Font(name = 'Arial')
                ws.cell(r, col_i).number_format = '#,##0'
            val_i += 1
    
    ws.column_dimensions.group(start='O', end='Q', hidden=True)
    ws.column_dimensions.group(start='V', end='X', hidden=True)

    wb.save(DIR_OUTPUT+'Summary.xlsx')
    return


def create_pivot():
    Excel = win32com.client.gencache.EnsureDispatch('Excel.Application') # Excel = win32com.client.Dispatch('Excel.Application')
    win32c = win32com.client.constants
    Excel.Visible = 0

    wb =Excel.Workbooks.Open(DIR_OUTPUT + str(datetime.now().strftime("%Y-%m-%d"))+' ADU Data.xlsx')
    Sheet1 = wb.Worksheets("Data")

    PivotSourceRange = Sheet1.UsedRange

    PivotSourceRange.Select()

    wb.Sheets.Add (After=wb.Sheets("Data"))
    Sheet2 = wb.Worksheets(2)
    Sheet2.Name = 'Pivot Table'
    cl3=Sheet2.Cells(1,1)
    PivotTargetRange=  Sheet2.Range(cl3,cl3)
    PivotTableName = 'ReportPivotTable'

    PivotCache = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=PivotSourceRange, Version=win32c.xlPivotTableVersion14)
    PivotTable = PivotCache.CreatePivotTable(TableDestination=PivotTargetRange, TableName=PivotTableName, DefaultVersion=win32c.xlPivotTableVersion14)

    PivotTable.PivotFields('ADU Ind').Orientation = win32c.xlRowField
    PivotTable.PivotFields('ADU Ind').Position = 1
    PivotTable.PivotFields('Year + Quarter').Orientation = win32c.xlRowField
    PivotTable.PivotFields('Year + Quarter').Position = 2
    PivotTable.PivotFields('In System').Orientation = win32c.xlPageField
    PivotTable.PivotFields('In System').Position = 1
    PivotTable.PivotFields('Primary Demo Non-ADU Equiv Deal Imp').Orientation = win32c.xlDataField
    PivotTable.PivotFields('Booked Dollars').Orientation = win32c.xlDataField
    PivotTable.PivotFields('Primary Demo Equiv Post Imp').Orientation = win32c.xlDataField
    PivotTable.PivotFields('Effec_Delv_Imp').Orientation = win32c.xlDataField
    PivotTable.PivotFields('Effec_Delv_value').Orientation = win32c.xlDataField
    PivotTable.PivotFields('Owed_Imp').Orientation = win32c.xlDataField
    PivotTable.PivotFields('Owed_value').Orientation = win32c.xlDataField

    
    # format
    
    ws = wb.Worksheets('Pivot Table')
    Sheet2.UsedRange.Font.Name = 'Arial'
    Sheet2.UsedRange.Font.Size = 10
    Sheet2.UsedRange.NumberFormat = "#,##0"
    
    PivotTable.InGridDropZones = True
    #Sheet2.PivotTables('ReportPivotTable').RowAxisLayout(win32c.xlTabularRow)
    PivotTable.RowAxisLayout(win32c.xlTabularRow)
    
    wb.Save()
    wb.Close(True)
    Excel.Application.Quit()   
    return 

def create_summary_pivot():
    Excel = win32com.client.gencache.EnsureDispatch('Excel.Application') 
    win32c = win32com.client.constants
    Excel.Visible = 1

    wb =Excel.Workbooks.Open(DIR_OUTPUT + str(datetime.now().strftime("%Y-%m-%d")) + ' Deal Delivery.xlsx')
    ws1 = wb.Worksheets("Deal Delivery")

    cl1 = ws1.Cells(7,3)
    cl2 = ws1.Cells(ws1.UsedRange.Rows.Count,65)
   
    PivotSourceRange = ws1.Range(cl1, cl2)
    PivotSourceRange.Select()

    wb.Sheets.Add(After=wb.Sheets("Deal Delivery"))
    ws2 = wb.Worksheets(2)
    ws2.Name = 'Pivot'
    
    # Pivot Table 1
    cl3=ws2.Cells(3,1)
    PivotTargetRange= ws2.Range(cl3,cl3)
    PivotTableName = 'SummaryPivotTable'

    PivotCache = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=PivotSourceRange, Version=win32c.xlPivotTableVersion14)
    PivotTable = PivotCache.CreatePivotTable(TableDestination=PivotTargetRange, TableName=PivotTableName, DefaultVersion=win32c.xlPivotTableVersion14)

    PivotTable.PivotFields('LinkedDeal').Orientation = win32c.xlRowField
    PivotTable.PivotFields('SOLD Deal Year').Orientation = win32c.xlColumnField
    #PivotTable.PivotFields('SOLD Guarantee ID').Orientation = win32c.xlDataField
    #PivotTable.PivotFields('Total Booked $').Orientation = win32c.xlDataField

    PivotTable.AddDataField(PivotTable.PivotFields('SOLD Guarantee ID'), "Num of Deals", win32c.xlCount)
    PivotTable.AddDataField(PivotTable.PivotFields('Total Booked $'))
 
    PivotTable.InGridDropZones = True
    PivotTable.RowAxisLayout(win32c.xlTabularRow)   


    # Pivot Table 2
    cl3=ws2.Cells(ws2.UsedRange.Rows.Count+7,1)
    PivotTargetRange= ws2.Range(cl3,cl3)
    PivotTableName = 'DemoTable'

    PivotCache = wb.PivotCaches().Create(SourceType=win32c.xlDatabase, SourceData=PivotSourceRange, Version=win32c.xlPivotTableVersion14)
    PT = PivotCache.CreatePivotTable(TableDestination=PivotTargetRange, TableName=PivotTableName, DefaultVersion=win32c.xlPivotTableVersion14)

    PT.PivotFields('SOLD Deal Year').Orientation = win32c.xlPageField
    PT.PivotFields('SOLD Primary Demo').Orientation = win32c.xlRowField
    #PT.PivotFields('ADUFlag').Orientation = win32c.xlColumnField

    PT.AddDataField(PT.PivotFields('SOLD Guarantee ID'), "Num of Deals", win32c.xlCount)
    PT.AddDataField(PT.PivotFields('Total Booked $'))
    PT.AddDataField(PT.PivotFields('Total Units'))
 
    PT.InGridDropZones = True
    PT.RowAxisLayout(win32c.xlTabularRow)   
    
    
    # format
    ws2 = wb.Worksheets('Pivot')
    ws2.UsedRange.Font.Name = 'Arial'
    ws2.UsedRange.Font.Size = 10
    ws2.UsedRange.NumberFormat = "#,##0"
    
   # Calculated Fields
    #PivotTable.CalculatedFields().Add('Dollar %','= Total Booked $ / GrossUnits')
    
    wb.Save()
    wb.Close(True)
    Excel.Application.Quit()   
    return 


def combine_xlsx_files():
    
    f1 = DIR_OUTPUT + 'YM -- 1 ION ADU 3.0 (Arjun) -- ' + str(datetime.now().strftime("%Y-%m-%d")) + '.xlsx'
    f2 = DIR_OUTPUT + str(datetime.now().strftime("%Y-%m-%d")) + ' ADUs to delete.xlsx'
    f3 = DIR_OUTPUT + str(datetime.now().strftime("%Y-%m-%d")) + ' Deal Delivery.xlsx'
    f4 = DIR_OUTPUT + 'Summary.xlsx'
    f5 = DIR_OUTPUT + str(datetime.now().strftime("%Y-%m-%d")) + ' Ratings Summary.xlsx'
    f6 = DIR_OUTPUT + str(datetime.now().strftime("%Y-%m-%d")) + ' ADU Notes.xlsx'
    f7 = DIR_OUTPUT + str(datetime.now().strftime("%Y-%m-%d")) + ' Liability Report.xlsx'

    wb_comb = xw.Book(f1)
    
    print('Combining ADUs to delete')
    wb2 = xw.Book(f2)
    ws2 = wb2.sheets('ADUs to delete')
    ws2.api.Copy(After=wb_comb.sheets("ADUs to schedule").api)
    wb2.close()

    print('Combining Deal Delivery')
    wb3 = xw.Book(f3)
    ws3 = wb3.sheets('Deal Delivery')
    ws3.api.Copy(Before=wb_comb.sheets("ADUs to schedule").api)
    ws4 = wb3.sheets('Pivot')
    ws4.api.Copy(After=wb_comb.sheets("ADUs to delete").api)
    wb3.close()
    
    print('Combining Ratings Summary')
    wb4 = xw.Book(f5)
    ws5 = wb4.sheets("Ratings Summary")
    ws5.api.Copy(After=wb_comb.sheets("ADUs to delete").api)
    wb4.close()
    
    print('Combining Summary')
    
    wb1 = xw.Book(f4)
    ws1 = wb1.sheets('Summary')
    ws1.api.Copy(Before=wb_comb.sheets("Deal Delivery").api)
    wb1.close()
    
    print('Saving file')
    wb_comb.save()
    wb_comb.close()


    print('Combining ADU notes and Quarterly Liability Report')
    wb_comb_notes = xw.Book(f6)
    wb7 = xw.Book(f7)
    ws7 = wb7.sheets('Liability Report')
    ws7.api.Copy(After = wb_comb_notes.sheets('ADU Notes').api)
    wb7.close()
    wb_comb_notes.save()
    wb_comb_notes.close()

    #wb_comb.app.quit()
    os.remove(f2)
    os.remove(f3)
    os.remove(f4)
    os.remove(f5)
    os.remove(f7)
    
    return

def copy_to_reports():
    file_names = [fn for fn in os.listdir(DIR_REPORT) if fn.startswith('YM -- 1 ION ADU 3.0 (Arjun) -- ')]
    for fn in file_names:
        os.remove(DIR_REPORT+fn)
    filename='YM -- 1 ION ADU 3.0 (Arjun) -- ' + str(datetime.now().strftime("%Y-%m-%d")) + '.xlsx'
    shutil.copy(DIR_OUTPUT + filename,DIR_REPORT + filename )


def forecast_actual(df, internal_estimates, four_q):
    C3_file = DIR_INPUT+'Quarterly C3 rating.csv'
    C3_rating = pd.read_csv(C3_file)
    actual_prev = get_ratings(df, C3_rating, int(four_q[0][0]), int(four_q[0][1]))
    actual_cur = get_ratings(df, C3_rating, int(four_q[1][0]), int(four_q[1][1]))
    
    forecast_prev = get_ratings(df, internal_estimates, int(four_q[0][0]), int(four_q[0][1]))
    forecast_cur = get_ratings(df, internal_estimates, int(four_q[1][0]), int(four_q[1][1]))
    forecast_next = get_ratings(df, internal_estimates, int(four_q[2][0]), int(four_q[2][1]))

    demo_sort_df = df.groupby(['Primary Demo']).sum()['Booked Dollars'].reset_index().sort_values('Booked Dollars', ascending=False)
    demo_sort = demo_sort_df['Primary Demo'].tolist()    
    
    all_df = [forecast_prev, forecast_cur, forecast_next, actual_prev, actual_cur]
    res = []
    for temp in all_df:
        temp = temp.set_index('Demo')
        temp = temp.reindex(demo_sort)
        temp = temp.reset_index()
        res.append(temp)
    return res

def liability_qtr_report(basic, liab_update):
    quar = sorted(liab_update['Year + Quarter'].unique())
    #liab_update['DealName_Linked'] = liab_update['Guarantee Name'].fillna(liab_update['Deal Name'])
    pvt = pd.pivot_table(liab_update[liab_update['In System']=='Y'], \
                         index = 'Guarantee ID',\
                         columns = 'Year + Quarter', \
                         values=['Primary Demo Equiv Post Imp', 'Primary Demo Non-ADU Equiv Deal Imp'], \
                         aggfunc=np.sum, fill_value=0, margins = True)

    total_delv_imp = 0
    total_guar_imp = 0
    for yq in quar:
        pvt[yq + ' Delv %'] = pvt['Primary Demo Equiv Post Imp'][yq] / pvt['Primary Demo Non-ADU Equiv Deal Imp'][yq]
        total_delv_imp += pvt['Primary Demo Equiv Post Imp'][yq]
        total_guar_imp += pvt['Primary Demo Non-ADU Equiv Deal Imp'][yq]
        
    pvt['Total Delv %'] = total_delv_imp / total_guar_imp
    pvt = pvt.replace([np.inf, -np.inf], 'ADU Only')
    
    pvt = pvt.reset_index()
    del pvt['Primary Demo Equiv Post Imp']
    del pvt['Primary Demo Non-ADU Equiv Deal Imp']
    
    pvt.columns = pvt.columns.get_level_values(0)
    pvt = pvt.fillna(0)
    
    basic_info = basic[['Guarantee ID', 'Deal ID','Guarantee Name', 'Deal Year', 'Advertiser','AE Name', 'Agency']]
    report_df = pd.merge(basic_info, pvt, on = 'Guarantee ID')
    
    return report_df 


def main(Q_num = 1):
    print("Reading Data")
    
    t1 = time.time()
    
    zf = zipfile.ZipFile(DIR_INPUT+'Dealmaker BI weekly reports.zip') 
    df = pd.read_csv(zf.open('Report 1.csv'))
    zf.close()

    date = datetime.now()+ dt.timedelta(days=7)
    date_string = str(date.strftime("%m/%d/%Y"))
    startdate = datetime.strptime(date_string, '%m/%d/%Y')
    
    quarters = pd.read_csv(DIR_INPUT+'timeList.csv')
    quarters['start_date'] = pd.to_datetime(quarters['start_date']).dt.strftime('%m/%d/%Y')
    four_q = find_quarters(quarters, startdate)
    # Find the start date of each quarter
    quarter_sd = quarter_startdate(quarters, four_q)
    startq = quarter_sd[1] #schedule start date
    endq = quarter_sd[1 + Q_num] # schedule end date
  
    # Read in Ratings
    ratings_file = DIR_INPUT+'Quarterly Internal Estimates.csv'    
    internal_estimates = pd.read_csv(ratings_file)
    ratings = get_ratings(df, internal_estimates, int(four_q[1][0]), int(four_q[1][1]))
    t2 = time.time()
    print('Time for reading files: ', t2 - t1)

    print("Keep a copy of raw zip file")
    copy_rename('Dealmaker BI weekly reports.zip', str(datetime.now().strftime("%Y-%m-%d")) + ' Dealmaker BI weekly reports.zip')

    print('Scheduling ADU and generating new data')
    raw = raw_result(df, quarters, date_string, startdate, ratings, four_q, startq, endq)
    general = new_data(raw, quarters)
    new = newdata_to_df(df, general, raw[1])
    t3 = time.time()
    print('Time for scheduling ADU and generating new data: ', t3 - t2)

    print('Calculating liability')
    liab = liability(new)
    liab_update = calc_units(liab, raw)
    report_df = liability_qtr_report(raw[1], liab_update)

    t4 = time.time()
    print('Time for computing liability: ', t4 - t3)

    print('Exporting ADU schedule file')
    sep = seperate(raw)
    format_df(sep[0], liab_update)
    format_take_back(sep[1], liab_update)
    format_cur_standing(raw[:2], liab_update)
    format_ADU_notes(sep[0])
    format_liability_qtr_report(raw[0], report_df)

    t5 = time.time()
    print('Time for exporting ADU schedule: ', t5 - t4)

    print('Creating pivot table')
    create_pivot()
    create_summary_pivot()
    t6 = time.time()
    print('Time for creating pivot table: ', t6-t5)

    print('Generating summary')
    quar, report_values = get_report_values(liab_update, four_q)
    get_summary(report_values, date_string, quar)
    t7 = time.time()
    print('Time for generating summary: ', t7-t6)
    
    print('Comparing forcast and actual ratings')
    all_ratings = forecast_actual(df, internal_estimates, four_q)
    format_forecast_actual(all_ratings, four_q)
    t8 = time.time()
    print('Time for comparing forcast and actual ratings: ', t8-t7)
    
    combine_xlsx_files()
    print('Done')

    copy_to_reports()
    print('Total Time: ', t8 - t1)
    return

main()